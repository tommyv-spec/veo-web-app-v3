# -*- coding: utf-8 -*-
"""
Background Worker for Veo Web App

######################################################################
# WORKER FILE VERSION: 2025-01-23-LATE-R2-FIX-V2
# LAST MODIFIED: 2025-01-23
# CHANGES: Added late R2 recovery for race conditions in redo/job flow
######################################################################

Handles:
- Job queue processing
- Progress updates
- Error recovery
- Graceful shutdown
"""

import json
import os
import threading
import time
import subprocess
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional, List, Set, Union
from concurrent.futures import ThreadPoolExecutor, as_completed
from queue import Queue, Empty
import traceback

from config import (
    JobStatus, ClipStatus, VideoConfig, APIKeysConfig, 
    DialogueLine, app_config, get_gemini_keys_from_env, get_openai_key_from_env,
    api_keys_config  # Global singleton for persistent key blocking
)
from models import (
    get_db, Job, Clip, JobLog, BlacklistEntry, GenerationLog,
    add_job_log, update_job_progress
)
from veo_generator import VeoGenerator, list_images, GENAI_AVAILABLE, describe_subject_for_continuity
from error_handler import VeoError, error_handler


# ============================================================
# WORKER VERSION - Update this on each deployment for tracking
# ============================================================
WORKER_VERSION = "ZIP-14-REDO-R2-RECOVERY-ON"
WORKER_TYPE = "api"  # This is the API worker (not flow/local)


def safe_images_dir(images_dir: Union[str, None]) -> Union[Path, None]:
    """
    Safely convert images_dir to Path, returning None for empty/blank strings.
    
    CRITICAL: Never call Path() on empty strings!
    Path("") becomes Path(".") which searches the current directory,
    leading to "No images found in ." errors for Flow jobs.
    
    Flow jobs have images_dir="" because frames are in R2/Flow storage, not local disk.
    """
    if not images_dir or not images_dir.strip():
        return None
    # Also catch if someone passes "." directly
    if images_dir.strip() == "." or images_dir.strip() == "..":
        return None
    return Path(images_dir)


def ensure_frames_present(job, images_dir: Path, db, add_job_log_func):
    """
    Ensure frames are present locally, recovering from R2 if needed.
    Call this BEFORE any filesystem access in redo to guarantee frames exist.
    
    Returns True if frames are present, raises RuntimeError if recovery fails.
    """
    exts = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp"}
    
    # If dir exists and has images, we're good
    if images_dir.exists():
        try:
            if any(p.suffix.lower() in exts for p in images_dir.iterdir()):
                return True
        except FileNotFoundError:
            # Directory was deleted between exists() and iterdir() - continue to R2 recovery
            print(f"[Redo] ensure_frames_present: directory existed but was deleted during check", flush=True)
        except Exception as e:
            # Other error during directory listing - continue to R2 recovery
            print(f"[Redo] ensure_frames_present: error listing directory: {e}", flush=True)
    
    # Force recovery attempt
    print(f"[Redo] ensure_frames_present: missing/empty -> R2 recovery for job {job.id[:8]}", flush=True)
    add_job_log_func(db, job.id, "[Redo] Local frames missing/empty. Attempting cloud recovery.", "WARNING", "redo")
    db.commit()
    
    # Parse keys (support both JSON and python-dict-string)
    frames_r2_keys = None
    raw = job.frames_storage_keys
    if raw:
        try:
            frames_r2_keys = json.loads(raw)
        except Exception:
            # Fallback for legacy rows
            import ast
            try:
                frames_r2_keys = ast.literal_eval(raw)
            except Exception:
                frames_r2_keys = None
    
    if not frames_r2_keys:
        add_job_log_func(
            db, job.id, 
            "⚠️ Redo failed: Original images were deleted and no cloud backup exists. "
            "Please create a new job with re-uploaded images.", 
            "ERROR", "redo"
        )
        db.commit()
        raise RuntimeError(
            "Original images unavailable. Cloud storage backup was not configured when this job was created. "
            "Please create a new job with re-uploaded images."
        )
    
    from backends.storage import is_storage_configured, get_storage
    if not is_storage_configured():
        add_job_log_func(
            db, job.id, 
            "⚠️ Redo failed: Cloud storage is not configured on this server. "
            "Cannot recover original images.", 
            "ERROR", "redo"
        )
        db.commit()
        raise RuntimeError(
            "Cloud storage is not configured on this server. Cannot recover original images. "
            "Please contact support or create a new job with re-uploaded images."
        )
    
    storage = get_storage()
    
    # Create directory if it doesn't exist
    try:
        images_dir.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        add_job_log_func(db, job.id, f"⚠️ Redo failed: Cannot create images directory: {e}", "ERROR", "redo")
        db.commit()
        raise RuntimeError(f"Cannot create images directory: {e}")
    
    ok = 0
    errors = []
    for filename, r2_key in frames_r2_keys.items():
        try:
            local_path = images_dir / filename
            storage.download_file(r2_key, local_path)
            ok += 1
        except Exception as e:
            errors.append(f"{filename}: {e}")
            print(f"[Redo] Failed to download {filename}: {e}", flush=True)
    
    if ok == 0:
        error_details = "; ".join(errors[:3])  # Show first 3 errors
        if len(errors) > 3:
            error_details += f" (and {len(errors) - 3} more)"
        add_job_log_func(
            db, job.id, 
            f"⚠️ Redo failed: Could not download any frames from cloud storage. Errors: {error_details}", 
            "ERROR", "redo"
        )
        db.commit()
        raise RuntimeError(f"Cloud recovery failed - could not download any frames. First error: {errors[0] if errors else 'Unknown'}")
    
    print(f"[Redo] ensure_frames_present: recovered {ok} frames", flush=True)
    add_job_log_func(db, job.id, f"✓ Recovered {ok} frames from cloud storage", "INFO", "redo")
    db.commit()
    return True


def is_flow_job(job) -> bool:
    """
    Check if a job is a Flow backend job.
    Returns True if the job should be handled by Flow worker, not API worker.
    """
    if not job:
        return False
    
    backend_value = getattr(job, 'backend', None)
    backend_str = str(backend_value).lower() if backend_value else ''
    has_flow_url = bool(getattr(job, 'flow_project_url', None))
    
    return (
        'flow' in backend_str or 
        backend_value == 'flow' or
        str(backend_value) == 'BackendType.FLOW' or
        has_flow_url
    )


class JobPausedException(Exception):
    """Exception raised when job is paused (not an error)"""
    pass


# Email notification settings
EMAIL_ALERTS_ENABLED = True
ALERT_EMAIL_TO = "kaveno.biz@gmail.com"
# Gmail SMTP settings - requires App Password (not regular password)
# Go to Google Account > Security > 2-Step Verification > App passwords
SMTP_SERVER = os.environ.get("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_EMAIL = os.environ.get("SMTP_EMAIL", "")  # Your Gmail address
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD", "")  # Gmail App Password

# Track sent alerts to avoid spam
_alerts_sent = {
    "low_keys_10": False,
    "no_keys": False,
}

def send_key_alert_email(alert_type: str, available_keys: int, total_keys: int = 0, job_id: str = None):
    """Send email alert when API keys are running low or exhausted."""
    global _alerts_sent
    
    if not EMAIL_ALERTS_ENABLED:
        return
    
    # Skip low_keys_10 alert - it's not critical and causes delays
    if alert_type == "low_keys_10":
        return
    
    # Check if we already sent this alert (reset when keys recover)
    if _alerts_sent.get(alert_type, False):
        return
    
    if not SMTP_EMAIL or not SMTP_PASSWORD:
        print(f"[Worker] ⚠️ EMAIL ALERT ({alert_type}): {available_keys} keys remaining", flush=True)
        _alerts_sent[alert_type] = True
        return
    
    # Run email sending in background thread to avoid blocking
    def send_async():
        try:
            msg = MIMEMultipart()
            msg['From'] = SMTP_EMAIL
            msg['To'] = ALERT_EMAIL_TO
            
            if alert_type == "no_keys":
                msg['Subject'] = "🚨 URGENT: All Veo API Keys Exhausted!"
                body = f"""🚨 ALL API KEYS EXHAUSTED

Your Veo Web App has run out of available API keys.
- Total Keys: {total_keys}
- Job ID: {job_id or 'N/A'}
- Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
"""
            else:
                msg['Subject'] = f"Veo API Alert: {alert_type}"
                body = f"API Key Alert: {alert_type}\nAvailable: {available_keys}"
            
            msg.attach(MIMEText(body, 'plain'))
            
            # Add 5 second timeout
            with smtplib.SMTP(SMTP_SERVER, SMTP_PORT, timeout=5) as server:
                server.starttls()
                server.login(SMTP_EMAIL, SMTP_PASSWORD)
                server.send_message(msg)
            
            print(f"[Worker] ✉️ Email alert sent: {alert_type}", flush=True)
        except Exception as e:
            print(f"[Worker] Email alert failed (non-blocking): {e}", flush=True)
    
    _alerts_sent[alert_type] = True
    # Fire and forget - don't block the worker
    threading.Thread(target=send_async, daemon=True).start()

def reset_key_alerts():
    """Reset alert flags when keys recover."""
    global _alerts_sent
    _alerts_sent["low_keys_10"] = False
    _alerts_sent["no_keys"] = False
    print("[Worker] Key alert flags reset", flush=True)



def get_api_keys_with_fallback(api_keys_json: str = None) -> APIKeysConfig:
    """Get API keys - uses global singleton to persist blocked keys state."""
    global api_keys_config
    
    api_keys_data = json.loads(api_keys_json) if api_keys_json else {}
    gemini_keys = api_keys_data.get("gemini_keys", [])
    openai_key = api_keys_data.get("openai_key")
    
    # If job provides keys, update the global config (but keep blocked state)
    if gemini_keys:
        # Only update if different keys provided
        if gemini_keys != api_keys_config.gemini_api_keys:
            api_keys_config.gemini_api_keys = gemini_keys
            print(f"[Worker] Updated Gemini keys from job: {len(gemini_keys)} keys", flush=True)
    
    if openai_key:
        api_keys_config.openai_api_key = openai_key
    
    # Log current state using KeyPoolManager
    from config import key_pool
    pool_status = key_pool.get_pool_status_summary(api_keys_config)
    print(f"[Worker] API Keys: {pool_status['available']} available, {pool_status['rate_limited']} rate-limited, {pool_status['invalid']} invalid", flush=True)
    
    return api_keys_config

class JobWorker:
    """
    Background worker that processes video generation jobs.
    
    Features:
    - Configurable worker pool
    - Real-time progress updates
    - Graceful shutdown
    - Error recovery
    """
    
    def __init__(self, max_workers: int = 3):
        self.max_workers = max_workers
        self.executor: Optional[ThreadPoolExecutor] = None
        self.running_jobs: Dict[str, VeoGenerator] = {}
        self.job_queue: Queue = Queue()
        self.shutdown_event = threading.Event()
        self.worker_thread: Optional[threading.Thread] = None
        
        # SSE subscribers (job_id -> list of queues)
        self.subscribers: Dict[str, List[Queue]] = {}
        self.subscribers_lock = threading.Lock()
        
        # Track clips currently being processed for redo (to prevent duplicates)
        self._processing_redo_clips: set = set()
        self._redo_lock = threading.Lock()
    
    def start(self):
        """Start the worker"""
        if self.executor is not None:
            return
        
        # ============================================================
        # STARTUP BANNER - Verify which version is running
        # ============================================================
        print(f"", flush=True)
        print(f"╔══════════════════════════════════════════════════════════════╗", flush=True)
        print(f"║  VEO WEB APP - API WORKER STARTING                           ║", flush=True)
        print(f"║  Version: {WORKER_VERSION:<51}║", flush=True)
        print(f"║  Type: {WORKER_TYPE:<54}║", flush=True)
        print(f"║  Max Workers: {self.max_workers:<47}║", flush=True)
        print(f"║  Startup Time: {datetime.utcnow().isoformat():<44}║", flush=True)
        print(f"╚══════════════════════════════════════════════════════════════╝", flush=True)
        print(f"", flush=True)
        print(f"[Worker {WORKER_VERSION}] Flow jobs will be BLOCKED - only API backend jobs processed", flush=True)
        print(f"[Worker {WORKER_VERSION}] safe_images_dir() helper active - no more Path('.') errors", flush=True)
        print(f"", flush=True)
        
        self.shutdown_event.clear()
        self.executor = ThreadPoolExecutor(max_workers=self.max_workers)
        
        # Start job processor thread
        self.worker_thread = threading.Thread(target=self._process_jobs, daemon=True)
        self.worker_thread.start()
        
        print(f"[Worker {WORKER_VERSION}] Started with {self.max_workers} workers", flush=True)
    
    def stop(self):
        """Stop the worker gracefully"""
        print("[Worker] Shutting down...")
        self.shutdown_event.set()
        
        # Cancel all running jobs
        for job_id, generator in list(self.running_jobs.items()):
            if generator is not None:
                generator.cancel()
        
        if self.executor:
            self.executor.shutdown(wait=True)
            self.executor = None
        
        print("[Worker] Shutdown complete")
    
    def _process_jobs(self):
        """Main job processing loop"""
        last_status_log = time.time()
        last_resume_check = time.time()
        
        while not self.shutdown_event.is_set():
            try:
                # Check for pending jobs in database
                self._check_pending_jobs()
                
                # Check for redo requests
                self._check_redo_queue()
                
                # Periodically check if paused jobs can be resumed (rate limits may have expired)
                # Check every 30 seconds
                if time.time() - last_resume_check > 30:
                    last_resume_check = time.time()
                    self._resume_waiting_jobs()
                
                # Log key pool status every 60 seconds ONLY if there are jobs running
                if time.time() - last_status_log > 60:
                    last_status_log = time.time()
                    try:
                        from config import key_pool, api_keys_config
                        running_count = len(self.running_jobs)
                        # Only log if there are jobs running (reduces log clutter)
                        if running_count > 0 and key_pool and api_keys_config:
                            status = key_pool.get_pool_status_summary(api_keys_config)
                            print(f"[KeyPool] Status: {status['available']} available, {status['rate_limited']} rate-limited, {status['invalid']} invalid | {running_count} jobs running", flush=True)
                    except Exception:
                        pass
                
                time.sleep(app_config.worker_poll_interval)
            except Exception as e:
                print(f"[Worker] Error in job processor: {e}")
                traceback.print_exc()
                time.sleep(5)
    
    def _check_redo_queue(self):
        """Check for clips that need redo and process them.
        Redos run independently of main jobs - they don't count against capacity.
        This is the PRIMARY redo processor - it starts redos immediately.
        """
        # First check if any keys are available - no point starting redos if all keys rate-limited
        try:
            from config import key_pool, api_keys_config
            if key_pool and api_keys_config:
                status = key_pool.get_pool_status_summary(api_keys_config)
                if status["available"] == 0:
                    # No keys available - skip redo processing this cycle
                    return
        except Exception:
            pass  # If we can't check, proceed anyway
        
        with get_db() as db:
            # VERSION MARKER: v2025-01-22-REDO-SEPARATION
            # IMPORTANT: This worker ONLY processes 'redo_queued' status (API backend redos)
            # Flow backend redos use 'flow_redo_queued' and are handled ONLY by Flow worker
            # This separation guarantees Flow redos never get picked up by API worker
            
            # DEBUG: First check ALL redo clips regardless of backend
            all_redo_clips = db.query(Clip).filter(
                Clip.status == ClipStatus.REDO_QUEUED.value  # Only API redos, NOT flow_redo_queued
            ).all()
            
            if all_redo_clips:
                print(f"[Worker {WORKER_VERSION}] Found {len(all_redo_clips)} TOTAL redo_queued clips (before filtering):", flush=True)
                for c in all_redo_clips:
                    job = db.query(Job).filter(Job.id == c.job_id).first()
                    backend_val = job.backend if job else 'NO_JOB'
                    flow_url = job.flow_project_url if job else None
                    is_flow = is_flow_job(job) if job else False
                    print(f"[Worker] DEBUG: Redo clip {c.id} (clip_index={c.clip_index}) -> job {c.job_id[:8] if c.job_id else 'NONE'} backend='{backend_val}' is_flow={is_flow}", flush=True)
                    
                    # Log to UI for Flow jobs to confirm we're skipping them
                    if is_flow:
                        add_job_log(
                            db, c.job_id,
                            f"[{WORKER_VERSION}] Flow redo detected in queue - skipping for API worker (Flow worker should handle)",
                            "DEBUG", "system"
                        )
                        db.commit()
            
            # Get clips queued for redo - ONLY for API backend jobs
            # Flow backend jobs use flow_redo_queued status and are handled by the local Flow worker
            redo_clips = db.query(Clip).join(Job).filter(
                Clip.status == ClipStatus.REDO_QUEUED.value,  # NEVER includes flow_redo_queued
                Job.backend == 'api',  # Only process API backend redos
                Job.flow_project_url.is_(None)  # EXTRA CHECK: No Flow project URL
            ).order_by(Clip.id.asc()).limit(5).all()
            
            # SAFETY NET: Filter again in Python using is_flow_job helper
            safe_redo_clips = []
            for clip in redo_clips:
                job = db.query(Job).filter(Job.id == clip.job_id).first()
                if not job:
                    print(f"[Worker] SKIP redo clip {clip.id}: No job found", flush=True)
                    continue
                
                # Use helper function for reliable Flow detection
                if is_flow_job(job):
                    print(f"[Worker {WORKER_VERSION}] SKIP redo clip {clip.id}: Flow job detected", flush=True)
                    continue
                
                # Only accept explicitly API backend jobs
                backend_str = str(job.backend).lower() if job.backend else ''
                if backend_str != 'api':
                    print(f"[Worker {WORKER_VERSION}] SKIP redo clip {clip.id}: Unknown backend '{job.backend}' (not 'api')", flush=True)
                    continue
                
                safe_redo_clips.append(clip)
            
            redo_clips = safe_redo_clips
            if redo_clips:
                print(f"[Worker {WORKER_VERSION}] After filter, found {len(redo_clips)} API-backend redo clips", flush=True)
            
            for clip in redo_clips:
                # Check if this clip is already being processed - MUST hold lock during check AND add
                with self._redo_lock:
                    if clip.id in self._processing_redo_clips:
                        continue  # Skip - already being processed
                    
                    # Limit concurrent redos to prevent overload (but don't block on main jobs)
                    if len(self._processing_redo_clips) >= 3:
                        print(f"[Worker] Max concurrent redos (3) reached, will process clip {clip.clip_index + 1} next cycle", flush=True)
                        break  # Max 3 concurrent redos
                    
                    # Add to processing set BEFORE submitting to prevent race condition
                    self._processing_redo_clips.add(clip.id)
                
                # ATOMICALLY mark as generating — use UPDATE ... WHERE status='redo_queued'
                # This prevents multiple workers/threads from picking up the same clip
                try:
                    from sqlalchemy import update
                    result = db.execute(
                        update(Clip)
                        .where(Clip.id == clip.id, Clip.status == ClipStatus.REDO_QUEUED.value)
                        .values(status=ClipStatus.GENERATING.value, started_at=datetime.utcnow())
                    )
                    db.commit()
                    
                    if result.rowcount == 0:
                        # Another worker already claimed this clip
                        print(f"[Worker] Clip {clip.clip_index + 1} already claimed by another worker — skipping", flush=True)
                        with self._redo_lock:
                            self._processing_redo_clips.discard(clip.id)
                        continue
                except Exception as e:
                    print(f"[Worker] Failed to mark clip {clip.id} as generating: {e}", flush=True)
                    with self._redo_lock:
                        self._processing_redo_clips.discard(clip.id)
                    continue
                
                print(f"[Worker {WORKER_VERSION}] Found redo request for clip {clip.clip_index + 1} - starting immediately", flush=True)
                
                # Start redo - it creates its own generator, independent of the main job
                self._start_redo(clip.job_id, clip.id)
    
    def _start_redo(self, job_id: str, clip_id: int):
        """Start processing a single clip redo"""
        print(f"[Worker {WORKER_VERSION}] _start_redo called for clip {clip_id}, job {job_id[:8]}", flush=True)
        
        if self.executor is None:
            print(f"[Worker] _start_redo: No executor, aborting", flush=True)
            # Remove from processing set if we can't actually start
            with self._redo_lock:
                self._processing_redo_clips.discard(clip_id)
            return
        
        # SAFETY CHECK: Verify this is EXPLICITLY an API backend job before processing
        with get_db() as db:
            job = db.query(Job).filter(Job.id == job_id).first()
            if job:
                # Use helper function for reliable Flow detection
                if is_flow_job(job):
                    print(f"[Worker {WORKER_VERSION}] BLOCKED: Refusing Flow job {job_id[:8]} at _start_redo", flush=True)
                    add_job_log(
                        db, job_id,
                        f"⚠️ [{WORKER_VERSION}] API worker blocked redo at _start_redo - Flow job",
                        "WARNING", "system"
                    )
                    db.commit()
                    with self._redo_lock:
                        self._processing_redo_clips.discard(clip_id)
                    return
                
                # Also block if not explicitly API
                backend_str = str(job.backend).lower() if job.backend else ''
                if backend_str != 'api':
                    print(f"[Worker {WORKER_VERSION}] BLOCKED: Unknown backend '{job.backend}' at _start_redo for job {job_id[:8]}", flush=True)
                    add_job_log(
                        db, job_id,
                        f"⚠️ [{WORKER_VERSION}] API worker blocked redo at _start_redo - unknown backend '{job.backend}'",
                        "WARNING", "system"
                    )
                    db.commit()
                    with self._redo_lock:
                        self._processing_redo_clips.discard(clip_id)
                    return
                    
                print(f"[Worker {WORKER_VERSION}] _start_redo: job {job_id[:8]} is API backend, proceeding", flush=True)
            else:
                print(f"[Worker] _start_redo: Job {job_id[:8]} not found!", flush=True)
                with self._redo_lock:
                    self._processing_redo_clips.discard(clip_id)
                return
        
        print(f"[Worker {WORKER_VERSION}] Starting redo for clip {clip_id} (job {job_id[:8]})", flush=True)
        self.executor.submit(self._run_redo, job_id, clip_id)
    
    def _check_pending_jobs(self):
        """Check for and start pending jobs (API backend only)"""
        if len(self.running_jobs) >= self.max_workers:
            return
        
        with get_db() as db:
            # Get ALL pending jobs first, then filter in Python for reliability
            all_pending = db.query(Job).filter(
                Job.status == JobStatus.PENDING.value
            ).order_by(Job.created_at.asc()).limit(20).all()
            
            # Filter out Flow jobs in Python (more reliable than SQL functions)
            pending = []
            for job in all_pending:
                backend = getattr(job, 'backend', None)
                backend_str = str(backend).lower() if backend else 'api'
                
                if backend_str == 'flow':
                    print(f"[Worker] SKIPPING Flow job {job.id[:8]} in pending check", flush=True)
                    continue
                
                pending.append(job)
                if len(pending) >= (self.max_workers - len(self.running_jobs)):
                    break
            
            # Only log if there are jobs to process (reduces clutter)
            if pending:
                print(f"[Worker] Found {len(pending)} API jobs to process (filtered from {len(all_pending)} total)", flush=True)
                for j in pending:
                    backend_val = getattr(j, 'backend', 'N/A')
                    print(f"[Worker] Processing job {j.id[:8]} (backend={backend_val})", flush=True)
            
            for job in pending:
                if job.id not in self.running_jobs:
                    self._start_job(job.id)
    
    def _start_job(self, job_id: str):
        """Start processing a job"""
        if self.executor is None:
            return
        
        # Double-check backend before starting
        with get_db() as db:
            job = db.query(Job).filter(Job.id == job_id).first()
            if job:
                backend = getattr(job, 'backend', None)
                backend_lower = str(backend).lower() if backend else ''
                if backend_lower == 'flow':
                    print(f"[Worker] BLOCKED: Job {job_id[:8]} is Flow backend - NOT starting", flush=True)
                    return
                print(f"[Worker] Starting job {job_id[:8]} (backend={backend})", flush=True)
        
        # Add to running_jobs IMMEDIATELY to prevent race condition
        # Use a placeholder until the real generator is created
        self.running_jobs[job_id] = None  # Placeholder
        
        self.executor.submit(self._run_job, job_id)
    
    def _run_redo(self, job_id: str, clip_id: int):
        """Run a single clip redo"""
        generator = None
        
        try:
            # Clip is already in _processing_redo_clips (added by _check_redo_queue)
            # Just verify it's still there (defensive check)
            with self._redo_lock:
                if clip_id not in self._processing_redo_clips:
                    print(f"[Worker] Clip {clip_id} not in processing set - may have been cancelled", flush=True)
                    return  # Now inside try block, so finally will run
            
            with get_db() as db:
                clip = db.query(Clip).filter(Clip.id == clip_id).first()
                job = db.query(Job).filter(Job.id == job_id).first()
                
                if not clip or not job:
                    print(f"[Worker] Clip {clip_id} or job {job_id} not found in database", flush=True)
                    return
                
                # Verify this clip is still ours (status should be GENERATING from atomic claim)
                if clip.status != ClipStatus.GENERATING.value:
                    print(f"[Worker] Clip {clip_id} status is '{clip.status}' not 'generating' — another worker owns it", flush=True)
                    return
                
                # ===== ULTRA VERBOSE DEBUG =====
                backend_raw = job.backend
                backend_str = str(backend_raw).lower() if backend_raw else ''
                images_dir_raw = job.images_dir
                flow_url = job.flow_project_url
                
                print(f"[Worker {WORKER_VERSION}] _run_redo DEBUG:", flush=True)
                print(f"  job.backend (raw) = '{backend_raw}'", flush=True)
                print(f"  job.backend (lower) = '{backend_str}'", flush=True)
                print(f"  job.images_dir = '{images_dir_raw}'", flush=True)
                print(f"  job.flow_project_url = '{flow_url}'", flush=True)
                print(f"  is_flow_job() = {is_flow_job(job)}", flush=True)
                
                add_job_log(
                    db, job_id,
                    f"[DEBUG] _run_redo: backend='{backend_raw}', images_dir='{images_dir_raw}', flow_url={bool(flow_url)}, is_flow={is_flow_job(job)}",
                    "DEBUG", "system"
                )
                db.commit()
                
                # ===== BLOCK 1: is_flow_job check =====
                if is_flow_job(job):
                    print(f"[Worker {WORKER_VERSION}] BLOCK 1: is_flow_job=True, refusing Flow job {job_id[:8]}", flush=True)
                    add_job_log(
                        db, job_id,
                        f"⚠️ [{WORKER_VERSION}] API worker blocked redo (BLOCK 1) - Flow job should be handled by Flow worker",
                        "WARNING", "system"
                    )
                    clip.status = ClipStatus.REDO_QUEUED.value
                    clip.error_message = None
                    db.commit()
                    return
                
                # ===== BLOCK 2: backend string check =====
                if backend_str == 'flow' or 'flow' in backend_str:
                    print(f"[Worker {WORKER_VERSION}] BLOCK 2: backend contains 'flow', refusing job {job_id[:8]}", flush=True)
                    add_job_log(
                        db, job_id,
                        f"⚠️ [{WORKER_VERSION}] API worker blocked redo (BLOCK 2) - backend='{backend_raw}'",
                        "WARNING", "system"
                    )
                    clip.status = ClipStatus.REDO_QUEUED.value
                    clip.error_message = None
                    db.commit()
                    return
                
                # ===== BLOCK 3: Not explicitly API =====
                if backend_str != 'api' and backend_str != '':
                    print(f"[Worker {WORKER_VERSION}] BLOCK 3: backend '{backend_str}' is not 'api', refusing job {job_id[:8]}", flush=True)
                    add_job_log(
                        db, job_id,
                        f"⚠️ [{WORKER_VERSION}] API worker blocked redo (BLOCK 3) - unknown backend '{backend_raw}'",
                        "WARNING", "system"
                    )
                    clip.status = ClipStatus.REDO_QUEUED.value
                    clip.error_message = None
                    db.commit()
                    return
                
                # ===== BLOCK 4: Has flow_project_url =====
                if flow_url:
                    print(f"[Worker {WORKER_VERSION}] BLOCK 4: has flow_project_url, refusing job {job_id[:8]}", flush=True)
                    add_job_log(
                        db, job_id,
                        f"⚠️ [{WORKER_VERSION}] API worker blocked redo (BLOCK 4) - has flow_project_url",
                        "WARNING", "system"
                    )
                    clip.status = ClipStatus.REDO_QUEUED.value
                    clip.error_message = None
                    db.commit()
                    return
                
                # Double-check status - if not REDO_QUEUED or GENERATING, someone else processed it
                if clip.status not in [ClipStatus.REDO_QUEUED.value, ClipStatus.GENERATING.value]:
                    print(f"[Worker] Clip {clip_id} status is {clip.status}, not REDO_QUEUED/GENERATING - skipping", flush=True)
                    return
                
                # ===== BLOCK 5: images_dir check =====
                images_dir = safe_images_dir(job.images_dir)
                print(f"[Worker {WORKER_VERSION}] safe_images_dir('{images_dir_raw}') returned: {images_dir} (type={type(images_dir)})", flush=True)
                
                if images_dir is None:
                    print(f"[Worker {WORKER_VERSION}] BLOCK 5: images_dir is None (empty), refusing job {job_id[:8]}", flush=True)
                    add_job_log(
                        db, job_id,
                        f"⚠️ [{WORKER_VERSION}] API worker blocked redo (BLOCK 5) - empty images_dir (likely Flow job)",
                        "WARNING", "system"
                    )
                    clip.status = ClipStatus.REDO_QUEUED.value
                    clip.error_message = "Empty images_dir - misrouted job"
                    db.commit()
                    return
                
                # ===== BLOCK 6: images_dir is "." =====
                if str(images_dir) in (".", "..", ""):
                    print(f"[Worker {WORKER_VERSION}] BLOCK 6: images_dir is '{images_dir}', refusing job {job_id[:8]}", flush=True)
                    add_job_log(
                        db, job_id,
                        f"⚠️ [{WORKER_VERSION}] API worker blocked redo (BLOCK 6) - invalid images_dir '{images_dir}'",
                        "WARNING", "system"
                    )
                    clip.status = ClipStatus.REDO_QUEUED.value
                    clip.error_message = None
                    db.commit()
                    return
                
                # ===== ENSURE FRAMES ARE PRESENT (R2 recovery if needed) =====
                # This single call handles all recovery logic and logs appropriately
                # We retry up to 2 times if there are transient errors (NOT re-queue, which causes race conditions)
                max_recovery_attempts = 2
                for recovery_attempt in range(max_recovery_attempts):
                    try:
                        ensure_frames_present(job, images_dir, db, add_job_log)
                        break  # Success - exit retry loop
                    except RuntimeError as recovery_err:
                        # Permanent failure (no R2 keys, storage not configured, etc.)
                        add_job_log(db, job_id, f"⚠️ Redo failed: {recovery_err}", "ERROR", "redo")
                        clip.status = ClipStatus.FAILED.value
                        clip.error_message = str(recovery_err)
                        db.commit()
                        return
                    except Exception as recovery_exc:
                        # Transient error - retry within same thread (NOT re-queue)
                        if recovery_attempt < max_recovery_attempts - 1:
                            print(f"[Worker] ensure_frames_present failed ({recovery_exc}), retrying ({recovery_attempt + 1}/{max_recovery_attempts})...", flush=True)
                            time.sleep(1)  # Brief pause before retry
                            continue
                        else:
                            # All retries exhausted - fail the clip
                            add_job_log(db, job_id, f"⚠️ Redo failed after {max_recovery_attempts} attempts: {recovery_exc}", "ERROR", "redo")
                            clip.status = ClipStatus.FAILED.value
                            clip.error_message = f"Recovery failed after {max_recovery_attempts} attempts: {str(recovery_exc)}"
                            db.commit()
                            return
                
                # Status already set to GENERATING by _check_redo_queue
                # Just log the redo start
                
                # DEBUG: Log frames_storage_keys status to diagnose R2 recovery issues
                frames_keys_status = "SET" if job.frames_storage_keys else "NULL"
                add_job_log(
                    db, job_id, 
                    f"[{WORKER_VERSION}] Starting redo for clip {clip.clip_index + 1} (attempt {clip.generation_attempt}/3) [R2 keys: {frames_keys_status}]",
                    "INFO", "redo"
                )
                
                # Parse configuration
                config_data = json.loads(job.config_json)
                config = VideoConfig(
                    aspect_ratio=config_data.get("aspect_ratio", "9:16"),
                    resolution=config_data.get("resolution", "720p"),
                    duration=config_data.get("duration", "8"),
                    language=config_data.get("language", "English"),
                    use_interpolation=config_data.get("use_interpolation", True),
                    use_openai_prompt_tuning=config_data.get("use_openai_prompt_tuning", True),
                    use_frame_vision=config_data.get("use_frame_vision", True),
                    max_retries_per_clip=config_data.get("max_retries_per_clip", 5),
                    custom_prompt=config_data.get("custom_prompt", ""),
                    user_context=config_data.get("user_context", ""),
                    single_image_mode=config_data.get("single_image_mode", False),
                    generation_mode=config_data.get("generation_mode", "parallel"),
                )
                
                # Parse API keys (with env fallback)
                api_keys = get_api_keys_with_fallback(job.api_keys_json)
                
                # SAFETY: Triple-check images_dir before using
                # This catches cases where old code or race conditions allowed bad values through
                if images_dir is None or str(images_dir) in ("", ".", ".."):
                    print(f"[Worker {WORKER_VERSION}] SAFETY BLOCK: images_dir={images_dir} is invalid for redo", flush=True)
                    add_job_log(
                        db, job_id,
                        f"⚠️ Redo blocked: invalid images_dir='{images_dir}'. This appears to be a Flow job - use Flow worker.",
                        "WARNING", "system"
                    )
                    clip.status = ClipStatus.REDO_QUEUED.value
                    clip.error_message = None
                    db.commit()
                    return
                
                output_dir = Path(job.output_dir)
                
                # FINAL SAFETY: Catch any edge case where images_dir is still invalid
                # This protects against old deployed code or race conditions
                if images_dir is None or not images_dir or str(images_dir).strip() in ("", ".", ".."):
                    print(f"[Worker {WORKER_VERSION}] FINAL SAFETY: images_dir='{images_dir}' invalid - likely Flow job", flush=True)
                    add_job_log(
                        db, job_id,
                        f"⚠️ [{WORKER_VERSION}] Redo skipped: Flow job detected (images_dir='{images_dir}'). Local Flow worker will handle.",
                        "WARNING", "system"
                    )
                    clip.status = ClipStatus.REDO_QUEUED.value
                    clip.error_message = None
                    db.commit()
                    return
                
                # Try to list images - with R2 recovery fallback for race conditions
                # The R2 check above may have passed but the directory could be cleared between check and list
                images = None
                try:
                    images = list_images(images_dir, config)
                except (FileNotFoundError, ValueError) as list_error:
                    # Directory was cleared between R2 check and list_images call - try R2 recovery now
                    print(f"[Worker {WORKER_VERSION}] list_images failed ({list_error}), attempting R2 recovery...", flush=True)
                    add_job_log(db, job_id, f"[Redo] Images dir missing at list_images, attempting R2 recovery", "WARNING", "redo")
                    db.commit()
                    
                    # Attempt R2 recovery (same logic as above but forced)
                    frames_r2_keys = None
                    if job.frames_storage_keys:
                        try:
                            frames_r2_keys = json.loads(job.frames_storage_keys)
                        except:
                            pass
                    
                    if frames_r2_keys:
                        try:
                            from backends.storage import is_storage_configured, get_storage
                            if is_storage_configured():
                                storage = get_storage()
                                
                                # Create local directory
                                images_dir.mkdir(parents=True, exist_ok=True)
                                
                                # Download all frames
                                downloaded_count = 0
                                for filename, r2_key in frames_r2_keys.items():
                                    try:
                                        local_path = images_dir / filename
                                        storage.download_file(r2_key, local_path)
                                        downloaded_count += 1
                                    except Exception as e:
                                        print(f"[Worker] Failed to download {filename}: {e}", flush=True)
                                
                                if downloaded_count > 0:
                                    print(f"[Worker {WORKER_VERSION}] Late R2 recovery: downloaded {downloaded_count} frames", flush=True)
                                    add_job_log(db, job_id, f"✓ Late R2 recovery: downloaded {downloaded_count} frames", "INFO", "redo")
                                    db.commit()
                                    # Retry list_images
                                    images = list_images(images_dir, config)
                                else:
                                    raise ValueError(f"R2 recovery downloaded 0 frames")
                            else:
                                raise ValueError(f"R2 storage not configured")
                        except Exception as r2_error:
                            add_job_log(db, job_id, f"⚠️ Late R2 recovery failed: {r2_error}", "ERROR", "redo")
                            clip.status = ClipStatus.FAILED.value
                            clip.error_message = f"Images unavailable and R2 recovery failed: {r2_error}"
                            db.commit()
                            return
                    else:
                        add_job_log(db, job_id, f"⚠️ Redo failed: No R2 backup available for recovery", "ERROR", "redo")
                        clip.status = ClipStatus.FAILED.value
                        clip.error_message = "Original images deleted and no cloud backup available. Please create a new job."
                        db.commit()
                        return
                
                if not images:
                    raise ValueError(f"No images found in {images_dir}")
                
                # Create generator for redo (uses dynamic key pool - all keys shared)
                generator = VeoGenerator(
                    config=config,
                    api_keys=api_keys,
                    openai_key=api_keys.openai_api_key,
                    job_id=job_id,
                )
                
                # Set up callbacks
                def on_progress(clip_index, status, message, details):
                    self._handle_progress(job_id, clip_index, status, message, details)
                
                def on_error(error):
                    self._handle_error(job_id, error)
                
                generator.on_progress = on_progress
                generator.on_error = on_error
                
                # Find frames from clip record
                start_frame = None
                end_frame = None
                start_index = 0
                end_index = 0
                
                # Log what we're looking for
                print(f"[Redo] Clip {clip.clip_index + 1}: Looking for start_frame='{clip.start_frame}', end_frame='{clip.end_frame}'", flush=True)
                
                # SAFEGUARD: Detect if stored frame name is an extracted frame (not original image)
                # Extracted frames have patterns like: "enhanced_", "_lastframe", "_extracted"
                def is_extracted_frame_name(name):
                    if not name:
                        return False
                    name_lower = name.lower()
                    return any(pattern in name_lower for pattern in ["enhanced_", "_lastframe", "_extracted", "lastframe_"])
                
                stored_start = clip.start_frame
                stored_end = clip.end_frame
                
                # If stored frame is an extracted frame, try to find the original scene image
                if is_extracted_frame_name(stored_start):
                    print(f"[Redo] WARNING: start_frame '{stored_start}' appears to be an extracted frame, will use scene image", flush=True)
                    # Try to extract original image name from the pattern (e.g., "1_image_03_20251222..." -> look for image_XX)
                    # Fall back to finding by clip's scene index
                    stored_start = None  # Will trigger fallback below
                
                for i, img in enumerate(images):
                    if stored_start and img.name == stored_start:
                        start_frame = img
                        start_index = i
                    if stored_end and img.name == stored_end:
                        end_frame = img
                        end_index = i
                
                if not start_frame:
                    # Fallback: Use original scene image based on clip index from dialogue
                    # Try to determine correct image from clip_index and scene structure
                    print(f"[Redo] WARNING: start_frame '{clip.start_frame}' not found in images", flush=True)
                    
                    # Parse scenes from config to find correct image for this clip
                    if config_data.get("dialogue_json"):
                        try:
                            dialogue_raw = json.loads(job.dialogue_json)
                            if isinstance(dialogue_raw, dict) and "lines" in dialogue_raw:
                                lines = dialogue_raw["lines"]
                                if clip.clip_index < len(lines):
                                    line_data = lines[clip.clip_index]
                                    scene_img_idx = line_data.get("start_image_idx", clip.clip_index % len(images))
                                    if scene_img_idx < len(images):
                                        start_frame = images[scene_img_idx]
                                        start_index = scene_img_idx
                                        print(f"[Redo] Using scene image from dialogue: {start_frame.name} at index {start_index}", flush=True)
                        except Exception as e:
                            print(f"[Redo] Failed to parse dialogue for scene image: {e}", flush=True)
                    
                    # Final fallback: use first image
                    if not start_frame:
                        print(f"[Redo] Using fallback: images[0] = {images[0].name}", flush=True)
                        start_frame = images[0]
                        start_index = 0
                else:
                    print(f"[Redo] Found start_frame: {start_frame.name} at index {start_index}", flush=True)
                
                # For interpolation: ONLY use end frame if the clip ORIGINALLY had one
                # CONTINUE mode clips have end_frame=None and should stay that way
                if clip.end_frame:
                    # Clip was created with an end frame - try to find it
                    if end_frame:
                        print(f"[Redo] Found end_frame: {end_frame.name} at index {end_index}", flush=True)
                    else:
                        # end_frame name is stored but file not found - try to find it again
                        print(f"[Redo] WARNING: end_frame '{clip.end_frame}' not found, searching again...", flush=True)
                        for i, img in enumerate(images):
                            if img.name == clip.end_frame:
                                end_frame = img
                                end_index = i
                                break
                        
                        if not end_frame:
                            # Still not found - use same image as fallback for interpolation
                            print(f"[Redo] end_frame still not found, using start_frame for interpolation", flush=True)
                            end_frame = start_frame
                            end_index = start_index
                else:
                    # Clip was created WITHOUT end frame (CONTINUE/FRESH mode)
                    # Keep it that way - no interpolation for this clip
                    print(f"[Redo] Clip has no end_frame (CONTINUE/FRESH mode) - no interpolation", flush=True)
                    end_frame = None
                    end_index = start_index
                
                print(f"[Redo] FINAL frames: start={start_frame.name if start_frame else None} (idx={start_index}), end={end_frame.name if end_frame else None} (idx={end_index})", flush=True)
                
                # Store original scene image for voice profile and subject description
                original_scene_image = start_frame
                
                # CONTINUE MODE: For clips that require previous clip's video frame
                # Determine clip_mode from job config (check line first, then scene)
                clip_mode = "blend"
                requires_previous = False
                
                try:
                    dialogue_raw = json.loads(job.dialogue_json)
                    scenes_data = None
                    dialogue_lines = []
                    
                    if isinstance(dialogue_raw, dict):
                        dialogue_lines = dialogue_raw.get("lines", [])
                        scenes_data = dialogue_raw.get("scenes", None)
                    else:
                        dialogue_lines = dialogue_raw
                    
                    if clip.clip_index < len(dialogue_lines):
                        line_data = dialogue_lines[clip.clip_index]
                        scene_idx = line_data.get("scene_index", 0)
                        
                        # Check line first for clip_mode
                        clip_mode = line_data.get("clip_mode")
                        
                        # If not on line, look up from scenes_data
                        if not clip_mode and scenes_data:
                            for scene in scenes_data:
                                if scene.get("sceneIndex") == scene_idx or scene.get("scene_index") == scene_idx:
                                    clip_mode = scene.get("clipMode") or scene.get("mode", "blend")
                                    break
                        
                        # Default to blend if still not found
                        if not clip_mode:
                            clip_mode = "blend"
                        
                        # Check if this clip requires previous (same scene, continue mode, not first clip)
                        if clip_mode == "continue" and clip.clip_index > 0:
                            prev_line = dialogue_lines[clip.clip_index - 1]
                            if prev_line.get("scene_index", 0) == scene_idx:
                                requires_previous = True
                        
                        print(f"[Redo] Clip {clip.clip_index + 1}: mode={clip_mode}, requires_previous={requires_previous}", flush=True)
                except Exception as e:
                    print(f"[Redo] Could not parse clip_mode from config: {e}", flush=True)
                
                if clip_mode == "continue" and requires_previous and clip.clip_index > 0:
                    print(f"[Redo] CONTINUE mode clip - checking for previous clip's video", flush=True)
                    
                    # Find previous clip
                    prev_clip = db.query(Clip).filter(
                        Clip.job_id == job_id,
                        Clip.clip_index == clip.clip_index - 1
                    ).first()
                    
                    if prev_clip and prev_clip.approval_status == "approved" and prev_clip.output_filename:
                        # Get previous clip's video path
                        prev_video_path = output_dir / prev_clip.output_filename
                        print(f"[Redo] Previous clip {prev_clip.clip_index + 1} video: {prev_video_path}", flush=True)
                        
                        if prev_video_path.exists():
                            # Extract frame from previous video
                            try:
                                import cv2
                                cap = cv2.VideoCapture(str(prev_video_path))
                                total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
                                
                                # Get frame near the end (8 frames from end)
                                target_frame = max(0, total_frames - 8)
                                cap.set(cv2.CAP_PROP_POS_FRAMES, target_frame)
                                ret, frame = cap.read()
                                cap.release()
                                
                                if ret:
                                    # Save extracted frame
                                    extracted_path = output_dir / f"redo_{clip.clip_index}_extracted.jpg"
                                    cv2.imwrite(str(extracted_path), frame)
                                    start_frame = extracted_path
                                    print(f"[Redo] Extracted frame from previous video: {extracted_path.name}", flush=True)
                                    add_job_log(db, job_id, f"Redo using extracted frame from clip {prev_clip.clip_index + 1}", "INFO", "redo")
                                else:
                                    print(f"[Redo] Failed to extract frame from previous video", flush=True)
                            except Exception as e:
                                print(f"[Redo] Error extracting frame: {e}", flush=True)
                        else:
                            print(f"[Redo] Previous video not found: {prev_video_path}", flush=True)
                    else:
                        print(f"[Redo] Previous clip not approved or has no video", flush=True)
                
                # Initialize voice profile for consistency (use original scene image)
                voice_id = generator.initialize_voice_profile(original_scene_image)
                add_job_log(db, job_id, f"Voice Profile for redo: {voice_id}", "INFO", "voice")
                
                # Determine prompt to use
                prompt_text = None
                redo_feedback = clip.redo_reason  # Get user's feedback
                
                if clip.use_logged_params and clip.prompt_text:
                    prompt_text = clip.prompt_text
                    add_job_log(db, job_id, f"Using logged parameters for redo", "INFO", "redo")
                else:
                    add_job_log(db, job_id, f"Using fresh parameters for redo", "INFO", "redo")
                
                if redo_feedback:
                    add_job_log(db, job_id, f"User feedback for redo: {redo_feedback}", "INFO", "redo")
                
                # Store clip values before leaving db context
                clip_dialogue_text = clip.dialogue_text
                clip_dialogue_id = clip.dialogue_id
                clip_clip_index = clip.clip_index
                clip_generation_attempt = clip.generation_attempt
                clip_use_logged_params = clip.use_logged_params
                
                self._broadcast_event(job_id, {
                    "type": "redo_started",
                    "clip_id": clip_id,
                    "clip_index": clip_clip_index,
                    "attempt": clip_generation_attempt,
                    "use_logged_params": clip_use_logged_params,
                    "redo_feedback": redo_feedback,
                })
            
            # Generate clip (outside db context to avoid long transactions)
            # Use original_scene_image for scene_image (for voice profile, subject description)
            # Use start_frame for actual generation (may be extracted frame for CONTINUE mode)
            result = generator.generate_single_clip(
                start_frame=start_frame,
                end_frame=end_frame,
                dialogue_line=clip_dialogue_text,
                dialogue_id=clip_dialogue_id,
                clip_index=clip_clip_index,
                output_dir=output_dir,
                images_list=images,
                current_end_index=end_index,
                scene_image=original_scene_image,  # Original scene image for analysis
                redo_feedback=redo_feedback,  # Pass user's feedback
                generation_mode=config.generation_mode,  # Pass generation mode for blacklist scoping
            )
            
            # Update clip with result
            with get_db() as db:
                clip = db.query(Clip).filter(Clip.id == clip_id).first()
                
                if clip:
                    clip.completed_at = datetime.utcnow()
                    
                    if clip.started_at:
                        clip.duration_seconds = (clip.completed_at - clip.started_at).total_seconds()
                    
                    if result["success"]:
                        new_filename = result["output_path"].name if result["output_path"] else None
                        
                        # Add to versions history (avoid duplicates)
                        versions = json.loads(clip.versions_json) if clip.versions_json else []
                        existing_attempts = [v.get('attempt') for v in versions]
                        
                        if clip.generation_attempt not in existing_attempts:
                            versions.append({
                                "attempt": clip.generation_attempt,
                                "filename": new_filename,
                                "generated_at": datetime.utcnow().isoformat(),
                            })
                            clip.versions_json = json.dumps(versions)
                        
                        # Update current output and select new variant
                        # Use position in versions list (1-indexed), not attempt number
                        clip.status = ClipStatus.COMPLETED.value
                        clip.output_filename = new_filename
                        clip.selected_variant = len(versions)  # Position, not attempt number
                        clip.prompt_text = result.get("prompt_text")
                        clip.approval_status = "pending_review"  # Reset to pending review
                        clip.error_code = None
                        clip.error_message = None
                        
                        # NOTE: Do NOT update clip.start_frame or clip.end_frame during redo!
                        # The original frames should be preserved. The redo just generates
                        # a new version of the clip using the same frames.
                        
                        # Upload to R2 for persistence (API jobs)
                        if result.get("output_path"):
                            try:
                                from backends.storage import is_storage_configured, get_storage
                                if is_storage_configured():
                                    storage = get_storage()
                                    r2_key = f"jobs/{job_id}/outputs/{new_filename}"
                                    storage.upload_file(str(result["output_path"]), r2_key, content_type='video/mp4')
                                    output_url = storage.get_presigned_url(r2_key, expires_in=86400 * 7)
                                    clip.output_url = output_url
                                    # Update version entry with URL
                                    versions[-1]["url"] = output_url
                                    clip.versions_json = json.dumps(versions)
                                    print(f"[Worker] Uploaded redo clip {clip.clip_index} to R2: {r2_key}", flush=True)
                            except Exception as r2_err:
                                print(f"[Worker] R2 upload failed for redo clip {clip.clip_index} (non-fatal): {r2_err}", flush=True)
                        
                        add_job_log(
                            db, job_id,
                            f"Redo completed for clip {clip.clip_index + 1} (attempt {clip.generation_attempt}/3)",
                            "INFO", "redo"
                        )
                    else:
                        # Check if this is a "no keys" situation - should re-queue, not fail
                        if result.get("no_keys") or result.get("should_pause"):
                            # Re-queue the redo - it will be picked up when keys are available
                            clip.status = ClipStatus.REDO_QUEUED.value
                            # Don't increment attempt count - this wasn't a real attempt
                            add_job_log(
                                db, job_id,
                                f"Redo for clip {clip.clip_index + 1} re-queued: API keys temporarily unavailable",
                                "WARNING", "redo"
                            )
                        else:
                            # Real failure - mark as failed
                            clip.status = ClipStatus.FAILED.value
                            if result.get("error"):
                                clip.error_code = result["error"].code.value
                                clip.error_message = result["error"].message
                            
                            add_job_log(
                                db, job_id,
                                f"Redo failed for clip {clip.clip_index + 1}: {result.get('error', 'Unknown error')}",
                                "ERROR", "redo"
                            )
                    
                    db.commit()
                
                # Determine event type based on result
                if result["success"]:
                    event_type = "redo_completed"
                elif result.get("no_keys") or result.get("should_pause"):
                    event_type = "redo_requeued"
                else:
                    event_type = "redo_failed"
                
                self._broadcast_event(job_id, {
                    "type": event_type,
                    "clip_id": clip_id,
                    "clip_index": clip.clip_index,
                    "success": result["success"],
                    "requeued": result.get("no_keys") or result.get("should_pause"),
                    "attempt": clip.generation_attempt,
                    "output": result["output_path"].name if result.get("output_path") else None,
                })
                
        except Exception as e:
            error = error_handler.classify_exception(e, {"job_id": job_id, "clip_id": clip_id})
            
            with get_db() as db:
                clip = db.query(Clip).filter(Clip.id == clip_id).first()
                job = db.query(Job).filter(Job.id == job_id).first()
                
                if clip:
                    error_str = str(e).lower()
                    
                    # Check if this was a Flow job wrongly processed by API worker
                    is_flow_job_error = False
                    if job:
                        backend_str = str(job.backend).lower() if job.backend else ''
                        images_dir_str = str(job.images_dir).strip() if job.images_dir else ''
                        # Flow jobs have: backend='flow', OR flow_project_url set, OR empty images_dir
                        is_flow_job_error = (
                            'flow' in backend_str or 
                            bool(job.flow_project_url) or
                            images_dir_str in ('', '.', '..')  # Empty images_dir = Flow job
                        )
                    
                    is_file_not_found = (
                        'file not found' in error_str or 
                        'no such file or directory' in error_str or
                        'errno 2' in error_str or
                        'no images found in .' in error_str or  # Flow job with empty images_dir
                        'images directory does not exist' in error_str or  # Missing dir
                        'images directory was deleted' in error_str or  # Race condition - dir deleted during access
                        'original images unavailable' in error_str or  # Recovery failed
                        'cannot access images directory' in error_str  # Permission or other OS error
                    )
                    
                    # DEBUG: Log the decision factors (console only, not UI)
                    print(f"[Worker] EXCEPTION DEBUG: job_id={job_id[:8]}, error_str[:100]='{error_str[:100]}'", flush=True)
                    print(f"[Worker] EXCEPTION DEBUG: job={job is not None}, backend={job.backend if job else 'N/A'}, is_flow_job_error={is_flow_job_error}, is_file_not_found={is_file_not_found}", flush=True)
                    
                    is_rate_limit = (
                        error.code.value == "RATE_LIMIT_429" or
                        "rate" in error_str or
                        "429" in error_str or
                        "no api keys available" in error_str or
                        "all keys are rate-limited" in error_str
                    )
                    
                    # If Flow job with file error, silently re-queue (don't log error - it's expected)
                    if is_flow_job_error and is_file_not_found:
                        clip.status = ClipStatus.REDO_QUEUED.value
                        clip.error_message = None  # Clear error so Flow worker can retry
                        clip.error_code = None
                        # Log that we're doing silent re-queue (debug only)
                        add_job_log(
                            db, job_id,
                            f"[DEBUG] Silent re-queue: Flow job file error handled correctly",
                            "DEBUG", "system"
                        )
                        print(f"[Worker] Flow job {job_id[:8]} silently re-queued for Flow worker (file not found is expected)", flush=True)
                    elif is_file_not_found and not is_flow_job_error:
                        # API job with missing files - this is a failure
                        # DO NOT re-queue here - it causes race conditions with duplicate redo threads
                        # The ensure_frames_present call should have already handled R2 recovery
                        # If we're here, recovery already failed or something else went wrong
                        has_r2_backup = bool(job and job.frames_storage_keys)
                        
                        # Log what happened for debugging
                        if has_r2_backup:
                            # R2 backup exists but recovery still failed - something else is wrong
                            # Fail instead of re-queuing to avoid race condition
                            clip.status = ClipStatus.FAILED.value
                            clip.error_code = "RECOVERY_FAILED"
                            clip.error_message = "Cloud recovery was attempted but files still unavailable. Please try again."
                            add_job_log(
                                db, job_id,
                                f"⚠️ Redo failed: Recovery from cloud was attempted but files remain unavailable. Error: {str(e)[:100]}",
                                "ERROR", "redo"
                            )
                            print(f"[Worker] API job {job_id[:8]} file missing even after R2 recovery - failing clip (not re-queueing)", flush=True)
                        else:
                            # No R2 backup - this is a permanent failure
                            clip.status = ClipStatus.FAILED.value
                            clip.error_code = "FILE_NOT_FOUND"
                            clip.error_message = "Original images no longer available. Please create a new job with re-uploaded images."
                            add_job_log(
                                db, job_id,
                                f"⚠️ Redo failed: Original images were deleted and no cloud backup was available. "
                                f"Cloud storage may not have been configured when this job was created. "
                                f"Please create a new job with re-uploaded images.",
                                "ERROR", "redo"
                            )
                            print(f"[Worker] API job {job_id[:8]} redo failed - no R2 backup available", flush=True)
                    elif is_rate_limit:
                        # Log and re-queue
                        self._handle_error(job_id, error)
                        clip.status = ClipStatus.REDO_QUEUED.value
                        add_job_log(
                            db, job_id,
                            f"Redo for clip {clip.clip_index + 1} re-queued: {error.message}",
                            "WARNING", "redo"
                        )
                    else:
                        # Real failure - log it
                        self._handle_error(job_id, error)
                        clip.status = ClipStatus.FAILED.value
                        clip.error_code = error.code.value
                        clip.error_message = error.message
                    db.commit()
        finally:
            # Always remove clip from processing set
            with self._redo_lock:
                self._processing_redo_clips.discard(clip_id)
            # Cleanup generator (dynamic keys, so just clears state)
            if generator:
                generator.cleanup()
    
    def _run_job(self, job_id: str):
        """Run a single job"""
        generator = None
        
        try:
            # FIRST: Update status to RUNNING immediately to prevent re-pickup
            with get_db() as db:
                job = db.query(Job).filter(Job.id == job_id).first()
                if not job:
                    return
                
                # Check if this is a Flow backend job - skip it (Flow worker handles it)
                backend_value = getattr(job, 'backend', None)
                # Handle both string and enum values
                backend_str = str(backend_value).lower() if backend_value else ''
                
                print(f"[Worker] Job {job_id[:8]} backend check:", flush=True)
                print(f"[Worker]   raw value: {backend_value!r}", flush=True)
                print(f"[Worker]   type: {type(backend_value)}", flush=True)
                print(f"[Worker]   lower str: '{backend_str}'", flush=True)
                
                # More comprehensive Flow check
                is_flow = (
                    backend_str == 'flow' or 
                    backend_value == 'flow' or
                    str(backend_value) == 'BackendType.FLOW' or
                    'flow' in backend_str
                )
                
                if is_flow:
                    print(f"[Worker] ❌ BLOCKED: Job {job_id[:8]} is Flow backend - NOT processing!", flush=True)
                    # Remove from running_jobs if we added it
                    if job_id in self.running_jobs:
                        del self.running_jobs[job_id]
                    return
                
                print(f"[Worker] ✓ Job {job_id[:8]} is API backend - proceeding", flush=True)
                
                # Check if already running (another thread got it)
                if job.status != JobStatus.PENDING.value:
                    print(f"[Worker] Job {job_id[:8]} already {job.status}, skipping", flush=True)
                    if job_id in self.running_jobs:
                        del self.running_jobs[job_id]
                    return
                
                # Update status FIRST
                job.status = JobStatus.RUNNING.value
                job.started_at = datetime.utcnow()
                db.commit()
                
                add_job_log(db, job_id, "Job started", "INFO", "system")
            
            # THEN: Do all the setup work
            # Step 1: Get job data (quick DB operation)
            with get_db() as db:
                job = db.query(Job).filter(Job.id == job_id).first()
                if not job:
                    return
                
                # Parse configuration
                config_data = json.loads(job.config_json)
                api_keys_json = job.api_keys_json
                images_dir = job.images_dir
                output_dir_str = job.output_dir
                dialogue_json = job.dialogue_json
                
                print(f"[Worker] Job {job_id[:8]}: Raw config_data = {config_data}")
                print(f"[Worker] Job {job_id[:8]}: Language from config = {config_data.get('language')}")
            
            # Step 2: Setup config (no DB needed)
            storyboard_mode = config_data.get("storyboard_mode", False)
            
            config = VideoConfig(
                aspect_ratio=config_data.get("aspect_ratio", "9:16"),
                resolution=config_data.get("resolution", "720p"),
                duration=config_data.get("duration", "8"),
                language=config_data.get("language", "English"),
                use_interpolation=config_data.get("use_interpolation", True),
                use_openai_prompt_tuning=config_data.get("use_openai_prompt_tuning", True),
                use_frame_vision=config_data.get("use_frame_vision", True),
                max_retries_per_clip=config_data.get("max_retries_per_clip", 5),
                custom_prompt=config_data.get("custom_prompt", ""),
                user_context=config_data.get("user_context", ""),
                single_image_mode=config_data.get("single_image_mode", False),
                generation_mode=config_data.get("generation_mode", "parallel"),
                skip_on_celebrity_filter=storyboard_mode,
            )
            
            with get_db() as db:
                add_job_log(db, job_id, f"Language: {config.language}", "INFO", "config")
            
            # Step 3: Parse API keys (no DB needed)
            api_keys = get_api_keys_with_fallback(api_keys_json)
            
            # Step 4: Validate API keys - EXPENSIVE OPERATION (HTTP calls)
            # This is done OUTSIDE the DB block to release the connection
            validation_logs = []
            def validation_log(msg):
                validation_logs.append(msg)
                print(msg, flush=True)
            
            working_now, rate_limited_count, invalid_count = api_keys.validate_keys_with_api(log_callback=validation_log)
            
            # Step 5: Log validation results (quick DB operation)
            with get_db() as db:
                for msg in validation_logs:
                    add_job_log(db, job_id, msg, "INFO", "system")
                
                job = db.query(Job).filter(Job.id == job_id).first()
                if not job:
                    return
                
                # Check if we have any working keys
                if working_now == 0:
                    if rate_limited_count > 0:
                        # All keys rate-limited - pause job and tell user to wait
                        job.status = JobStatus.PAUSED.value
                        db.commit()
                        add_job_log(
                            db, job_id,
                            f"⏸️ Job paused: All {rate_limited_count} API keys are rate-limited (429). Wait 5 minutes or add new keys, then Resume.",
                            "WARNING", "system"
                        )
                        # Raise special exception that won't mark job as failed
                        raise JobPausedException(f"All API keys are rate-limited. Job paused.")
                    else:
                        # No valid keys at all
                        raise ValueError("No valid API keys available. All keys are suspended or invalid. Please add working API keys.")
                
                # Adjust parallel_clips based on working keys (max 6, min 1)
                original_parallel = config.parallel_clips
                config.parallel_clips = max(1, min(working_now, 6))
                
                add_job_log(
                    db, job_id, 
                    f"🔑 {working_now} working API keys available ({rate_limited_count} rate-limited). Running {config.parallel_clips} parallel clips",
                    "INFO", "system"
                )
                
                # COMMIT HERE to ensure the above log is saved before anything else happens
                db.commit()
                
                if config.parallel_clips != original_parallel:
                    print(f"[Worker] Adjusted parallel_clips: {original_parallel} → {config.parallel_clips}", flush=True)
                
                # EXPLICIT LOG: We're about to start the dangerous section
                add_job_log(db, job_id, "[DEBUG] Checkpoint A: After key validation, before dialogue parsing", "DEBUG", "system")
                db.commit()
                
                # DEBUG: Log each step to find where it fails
                add_job_log(db, job_id, "[DEBUG] Step 1: About to parse dialogue JSON", "DEBUG", "system")
                db.commit()
                
                print(f"[Worker] DEBUG: About to parse dialogue JSON...", flush=True)
                
                # Parse dialogue data (new format includes scenes)
                dialogue_raw = json.loads(dialogue_json)
                
                add_job_log(db, job_id, "[DEBUG] Step 2: Dialogue parsed OK", "DEBUG", "system")
                db.commit()
                
                print(f"[Worker] DEBUG: Dialogue parsed, checking format...", flush=True)
                
                # Handle both old format (list) and new format (dict with lines/scenes)
                if isinstance(dialogue_raw, list):
                    # Old format: just a list of lines
                    dialogue_data = dialogue_raw
                    scenes_data = None
                    last_frame_index = None
                else:
                    # New format: {lines: [...], scenes: [...], last_frame_index: ...}
                    dialogue_data = dialogue_raw.get("lines", [])
                    scenes_data = dialogue_raw.get("scenes", None)
                    last_frame_index = dialogue_raw.get("last_frame_index", None)
                
                print(f"[Worker] DEBUG: Dialogue format OK, storyboard_mode={storyboard_mode}", flush=True)
                
                # Store scenes data for processing
                print(f"[Worker] Storyboard mode: {storyboard_mode}", flush=True)
                if storyboard_mode:
                    print(f"[Worker] Celebrity filter will SKIP clips (not retry)", flush=True)
                if scenes_data:
                    print(f"[Worker] Scenes: {json.dumps(scenes_data, indent=2)}", flush=True)
                if last_frame_index is not None:
                    print(f"[Worker] Last frame index: {last_frame_index}", flush=True)
                
                # Get images - use safe_images_dir helper to prevent Path(".") bug
                images_dir_path = safe_images_dir(images_dir)
                
                # CRITICAL: API jobs MUST have a valid images_dir
                if images_dir_path is None:
                    raise ValueError(f"Job has empty images_dir - this should not happen for API jobs. Job may be misconfigured.")
                
                if not images_dir_path.exists():
                    # Local files missing - try to recover from R2 storage
                    print(f"[Worker {WORKER_VERSION}] Local images_dir missing, attempting R2 recovery...", flush=True)
                    
                    # Re-fetch job to get frames_storage_keys
                    job = db.query(Job).filter(Job.id == job_id).first()
                    frames_r2_keys = None
                    if job and job.frames_storage_keys:
                        try:
                            frames_r2_keys = json.loads(job.frames_storage_keys)
                        except:
                            pass
                    
                    if frames_r2_keys:
                        try:
                            from backends.storage import is_storage_configured, get_storage
                            if is_storage_configured():
                                storage = get_storage()
                                
                                # Create local directory
                                images_dir_path.mkdir(parents=True, exist_ok=True)
                                
                                # Download all frames
                                downloaded_count = 0
                                for filename, r2_key in frames_r2_keys.items():
                                    try:
                                        local_path = images_dir_path / filename
                                        storage.download_file(r2_key, local_path)
                                        downloaded_count += 1
                                    except Exception as e:
                                        print(f"[Worker] Failed to download {filename}: {e}", flush=True)
                                
                                if downloaded_count > 0:
                                    print(f"[Worker {WORKER_VERSION}] Recovered {downloaded_count} frames from R2", flush=True)
                                    add_job_log(db, job_id, f"✓ Recovered {downloaded_count} frames from cloud storage", "INFO", "system")
                                    db.commit()
                                else:
                                    raise ValueError(f"No frames could be downloaded from R2 storage")
                            else:
                                raise ValueError(f"R2 storage not configured and images_dir does not exist: {images_dir_path}")
                        except Exception as e:
                            raise ValueError(f"Images directory does not exist and R2 recovery failed: {images_dir_path} ({e})")
                    else:
                        raise ValueError(f"Images directory does not exist: {images_dir_path}")
                
                output_dir = Path(output_dir_str)
                output_dir.mkdir(parents=True, exist_ok=True)
                
                # Try to list images - with R2 recovery fallback for race conditions
                images = None
                try:
                    images = list_images(images_dir_path, config)
                except (FileNotFoundError, ValueError) as list_error:
                    # Directory was cleared between R2 check and list_images call - try R2 recovery now
                    print(f"[Worker {WORKER_VERSION}] list_images failed ({list_error}), attempting late R2 recovery...", flush=True)
                    add_job_log(db, job_id, f"Images dir missing at list_images, attempting late R2 recovery", "WARNING", "system")
                    db.commit()
                    
                    # Re-fetch R2 keys
                    job = db.query(Job).filter(Job.id == job_id).first()
                    frames_r2_keys = None
                    if job and job.frames_storage_keys:
                        try:
                            frames_r2_keys = json.loads(job.frames_storage_keys)
                        except:
                            pass
                    
                    if frames_r2_keys:
                        try:
                            from backends.storage import is_storage_configured, get_storage
                            if is_storage_configured():
                                storage = get_storage()
                                
                                # Create local directory
                                images_dir_path.mkdir(parents=True, exist_ok=True)
                                
                                # Download all frames
                                downloaded_count = 0
                                for filename, r2_key in frames_r2_keys.items():
                                    try:
                                        local_path = images_dir_path / filename
                                        storage.download_file(r2_key, local_path)
                                        downloaded_count += 1
                                    except Exception as e:
                                        print(f"[Worker] Failed to download {filename}: {e}", flush=True)
                                
                                if downloaded_count > 0:
                                    print(f"[Worker {WORKER_VERSION}] Late R2 recovery: downloaded {downloaded_count} frames", flush=True)
                                    add_job_log(db, job_id, f"✓ Late R2 recovery: downloaded {downloaded_count} frames", "INFO", "system")
                                    db.commit()
                                    # Retry list_images
                                    images = list_images(images_dir_path, config)
                                else:
                                    raise ValueError(f"Late R2 recovery downloaded 0 frames")
                            else:
                                raise ValueError(f"R2 storage not configured for late recovery")
                        except Exception as r2_error:
                            raise ValueError(f"Images directory unavailable and late R2 recovery failed: {r2_error}")
                    else:
                        raise ValueError(f"Images directory unavailable and no R2 backup: {images_dir_path}")
                
                if not images:
                    raise ValueError(f"No images found in {images_dir_path}")
                
                add_job_log(db, job_id, f"[DEBUG] Step 3: Loaded {len(images)} images", "DEBUG", "system")
                db.commit()
                
                # Log image order for debugging
                print(f"[Worker] Loaded {len(images)} images in order:", flush=True)
                for idx, img in enumerate(images):
                    print(f"  [{idx}] {img.name}", flush=True)
                
                # Create generator with job_id for key reservation
                add_job_log(db, job_id, "[DEBUG] Step 4: Creating VeoGenerator...", "DEBUG", "system")
                db.commit()
                
                generator = VeoGenerator(
                    config=config,
                    api_keys=api_keys,
                    openai_key=api_keys.openai_api_key,
                    job_id=job_id,
                )
                
                add_job_log(db, job_id, "[DEBUG] Step 5: VeoGenerator created OK", "DEBUG", "system")
                db.commit()
                
                # Check if ALL keys are rate-limited or invalid (using dynamic pool)
                from config import key_pool
                pool_status = key_pool.get_pool_status_summary(api_keys)
                if pool_status["available"] == 0:
                    # No keys available - pause job to wait for rate limits to clear
                    job.status = JobStatus.PAUSED.value
                    db.commit()
                    
                    add_job_log(
                        db, job_id,
                        f"⏸️ Job queued: All {pool_status['total']} API keys are rate-limited ({pool_status['rate_limited']}) or invalid ({pool_status['invalid']}). Will auto-resume when keys recover.",
                        "INFO", "system"
                    )
                    print(f"[Worker] Job {job_id[:8]} paused - {pool_status['rate_limited']} keys rate-limited, {pool_status['invalid']} invalid", flush=True)
                    raise JobPausedException("No API keys available - waiting for rate limits to clear")
                
                # Set up callbacks
                def on_progress(clip_index, status, message, details):
                    self._handle_progress(job_id, clip_index, status, message, details)
                
                def on_error(error: VeoError):
                    self._handle_error(job_id, error)
                
                generator.on_progress = on_progress
                generator.on_error = on_error
                
                # Initialize voice profile ONCE for entire job (use first image as reference)
                add_job_log(db, job_id, "[DEBUG] Step 6: About to initialize voice profile...", "DEBUG", "system")
                db.commit()
                
                voice_id = generator.initialize_voice_profile(images[0])
                add_job_log(db, job_id, f"Voice Profile initialized: {voice_id}", "INFO", "voice")
                add_job_log(db, job_id, "[DEBUG] Step 7: Voice profile OK, starting clips", "DEBUG", "system")
                db.commit()
                
                self.running_jobs[job_id] = generator
            
            # Process clips (pass scenes_data for storyboard mode)
            self._process_clips(job_id, generator, dialogue_data, images, output_dir, scenes_data, last_frame_index)
        
        except JobPausedException as e:
            # Job was paused intentionally - don't mark as failed
            print(f"[Worker] Job {job_id[:8]} paused: {e}", flush=True)
            # Status already set to PAUSED, just return
        
        except Exception as e:
            # Log the RAW exception before classification (helps debug misclassifications)
            import traceback
            raw_error = f"{type(e).__name__}: {str(e)[:300]}"
            tb_str = traceback.format_exc()[-1000:]
            print(f"[Worker] RAW EXCEPTION in job {job_id[:8]}:", flush=True)
            print(f"[Worker]   Type: {type(e).__name__}", flush=True)
            print(f"[Worker]   Message: {str(e)[:500]}", flush=True)
            print(f"[Worker]   Traceback: {tb_str}", flush=True)
            
            error = error_handler.classify_exception(e, {"job_id": job_id})
            self._handle_error(job_id, error)
            
            # ALWAYS log raw error first (separate DB transaction to guarantee it's saved)
            # Use multiple fallback attempts
            for attempt in range(3):
                try:
                    with get_db() as db:
                        add_job_log(
                            db, job_id,
                            f"[RAW ERROR attempt {attempt+1}] {raw_error}",
                            "ERROR", "system"
                        )
                        # Also log the traceback
                        add_job_log(
                            db, job_id,
                            f"[TRACEBACK] {tb_str[:500]}",
                            "ERROR", "system"
                        )
                        db.commit()
                        break  # Success, exit retry loop
                except Exception as log_err:
                    print(f"[Worker] Failed to log error (attempt {attempt+1}): {log_err}", flush=True)
                    if attempt == 2:
                        print(f"[Worker] GIVING UP on error logging for job {job_id[:8]}", flush=True)
            
            # Only mark as failed if no clips succeeded
            with get_db() as db:
                job = db.query(Job).filter(Job.id == job_id).first()
                if job and job.status != JobStatus.PAUSED.value:  # Don't override if paused
                    clips = db.query(Clip).filter(Clip.job_id == job_id).all()
                    successful = sum(1 for c in clips if c.status == ClipStatus.COMPLETED.value)
                    
                    if successful == 0:
                        # No clips succeeded - mark job as failed
                        job.status = JobStatus.FAILED.value
                    else:
                        # Some clips succeeded - mark as completed with failures
                        job.status = JobStatus.COMPLETED.value
                    
                    job.completed_at = datetime.utcnow()
                    
                    # Log the classified error
                    add_job_log(
                        db, job_id, 
                        f"Job ended with error: {error.message}", 
                        "ERROR", "system",
                        details=error.to_dict()
                    )
                    
                    db.commit()
        
        finally:
            if job_id in self.running_jobs:
                # Release keys back to pool
                generator = self.running_jobs.get(job_id)
                if generator:
                    generator.cleanup()
                del self.running_jobs[job_id]
            
            # Check for waiting jobs and resume them now that keys are free
            # BUT only if this job completed successfully (not paused due to rate limit)
            with get_db() as db:
                job = db.query(Job).filter(Job.id == job_id).first()
                if job and job.status != JobStatus.PAUSED.value:
                    # Job completed or failed (not paused) - keys are truly free
                    self._resume_waiting_jobs()
    
    def _resume_waiting_jobs(self):
        """Check for paused jobs waiting for keys and resume them.
        
        With the NEW dynamic key pool, we just check if any keys are available
        (not rate-limited). No need to check for "free" vs "reserved" keys since
        all keys are shared dynamically.
        
        NOTE: Only resumes API backend jobs. Flow jobs are handled by Flow worker.
        """
        try:
            from config import key_pool, api_keys_config
            
            # Check if any keys are actually available using KeyPoolManager
            if api_keys_config and key_pool:
                status = key_pool.get_pool_status_summary(api_keys_config)
                if status["available"] == 0:
                    print(f"[Worker] No keys available ({status['rate_limited']} rate-limited) - skipping auto-resume", flush=True)
                    return
                available_count = status["available"]
            else:
                return
            
            with get_db() as db:
                # Find paused jobs (waiting for keys) - ONLY API backend jobs
                from sqlalchemy import or_
                paused_jobs = db.query(Job).filter(
                    Job.status == JobStatus.PAUSED.value,
                    or_(Job.backend == "api", Job.backend == None)  # Only API jobs
                ).order_by(Job.created_at.asc()).all()  # FIFO order
                
                if not paused_jobs:
                    return
                
                print(f"[Worker] Found {len(paused_jobs)} paused API job(s), {available_count} keys available", flush=True)
                
                for job in paused_jobs:
                    # With dynamic keys, just resume if ANY keys are available
                    job.status = JobStatus.PENDING.value
                    db.commit()
                    self.job_queue.put(job.id)
                    add_job_log(
                        db, job.id,
                        f"▶️ Job auto-resumed: {available_count} API key(s) now available.",
                        "INFO", "system"
                    )
                    print(f"[Worker] Auto-resumed paused job {job.id[:8]} ({available_count} keys available)", flush=True)
                    # Only resume one job at a time to prevent overload
                    break
                        
        except Exception as e:
            print(f"[Worker] Error checking waiting jobs: {e}", flush=True)
    
    def _process_clips(
        self,
        job_id: str,
        generator: VeoGenerator,
        dialogue_data: List[Dict],
        images: List[Path],
        output_dir: Path,
        scenes_data: Optional[List[Dict]] = None,
        last_frame_index: Optional[int] = None,
    ):
        """Process all clips for a job - with parallel generation support and scene-aware sequencing"""
        from concurrent.futures import ThreadPoolExecutor, as_completed, wait, FIRST_COMPLETED
        import subprocess
        
        # Check for single image mode
        single_image_mode = getattr(generator.config, 'single_image_mode', False) or len(images) == 1
        
        images_dir_str = str(images[0].parent)
        
        total_clips = len(dialogue_data)
        completed = 0
        failed = 0
        skipped = 0
        
        # Get parallel clip count from config (default 1 for memory efficiency on free tier)
        parallel_clips = getattr(generator.config, 'parallel_clips', 1)
        
        # Key exhaustion tracking
        no_keys_retries = 0
        max_no_keys_retries = 3
        no_keys_wait_seconds = 300  # 5 minutes
        
        # Log last frame
        if last_frame_index is not None:
            print(f"[Worker] Last frame index set: {last_frame_index} ({images[last_frame_index].name if last_frame_index < len(images) else 'INVALID'})", flush=True)
        
        # === BUILD SCENE-AWARE CLIP STRUCTURE ===
        num_images = len(images)
        use_interpolation = getattr(generator.config, 'use_interpolation', True)
        
        print(f"[Worker] Processing {total_clips} clips with {num_images} images", flush=True)
        print(f"[Worker] Scenes data: {scenes_data}", flush=True)
        
        # Build clip info with scene awareness
        clip_info = []  # List of dicts with all clip metadata
        
        for i, line_data in enumerate(dialogue_data):
            # Determine clip_mode - check line first, then scene
            clip_mode = line_data.get("clip_mode")
            scene_idx = line_data.get("scene_index", 0)
            
            # If not on line, look up from scenes_data
            if not clip_mode and scenes_data:
                for scene in scenes_data:
                    if scene.get("sceneIndex") == scene_idx or scene.get("scene_index") == scene_idx:
                        clip_mode = scene.get("clipMode") or scene.get("mode", "blend")
                        break
            
            # Default to blend if still not found
            if not clip_mode:
                clip_mode = "blend"
            
            info = {
                "index": i,
                "text": line_data["text"],
                "dialogue_id": line_data["id"],
                "image_idx": line_data.get("start_image_idx", i % num_images) if not single_image_mode else 0,
                "scene_index": scene_idx,
                "clip_mode": clip_mode,
                "scene_transition": line_data.get("scene_transition"),  # 'blend' | 'cut' | None
                "requires_previous": False,  # Will be set below
                "start_frame": None,  # Will be set or calculated
                "end_frame": None,    # Will be set or calculated
            }
            
            # Determine if this clip requires the previous clip to complete first
            # This happens when clip_mode is 'continue' AND it's not the first clip in its scene
            if info["clip_mode"] == "continue" and i > 0:
                prev_scene = dialogue_data[i-1].get("scene_index", 0)
                if prev_scene == info["scene_index"]:
                    # Same scene, continue mode - must wait for previous clip
                    info["requires_previous"] = True
            
            clip_info.append(info)
            print(f"[Worker] Clip {i}: scene={info['scene_index']}, mode={info['clip_mode']}, requires_prev={info['requires_previous']}", flush=True)
        
        # Calculate initial frame assignments
        # 
        # FRAME ASSIGNMENT LOGIC:
        # 
        # For each clip, we need to determine:
        # 1. START FRAME: Where the clip begins
        # 2. END FRAME: Where the clip ends (can be None for no interpolation)
        #
        # The logic depends on clip_mode:
        #
        # BLEND mode (standard):
        #   - Start: assigned image
        #   - End: depends on NEXT clip (allows smooth transitions)
        #
        # CONTINUE mode:
        #   - Start: extracted from previous clip's last frame (set at runtime)
        #   - End: depends on NEXT clip (allows smooth transitions)
        #
        # FRESH mode:
        #   - Start: always original image
        #   - End: NONE (completely standalone clips, no interpolation)
        #
        # END FRAME determination (for BLEND and CONTINUE modes only):
        #   - If LAST clip of video:
        #     - If Last Frame defined: use Last Frame
        #     - Else: None (no interpolation)
        #   - If NEXT clip is in SAME scene:
        #     - None (no end frame, natural continuation)
        #   - If NEXT clip is in DIFFERENT scene:
        #     - If transition = "blend": next scene's image
        #     - If transition = "cut": None
        #
        with get_db() as db:
            for i, info in enumerate(clip_info):
                start_idx = info["image_idx"]
                clip_mode = info["clip_mode"]
                scene_transition = info["scene_transition"]
                scene_index = info["scene_index"]
                
                # Default start: our assigned image
                actual_start_idx = start_idx
                
                # Determine END FRAME based on what comes AFTER this clip
                use_end_frame = False
                actual_end_idx = None
                end_frame_reason = ""
                
                # SINGLE IMAGE MODE: Always use same image as end frame for interpolation
                if single_image_mode and generator.config.use_interpolation:
                    use_end_frame = True
                    actual_end_idx = start_idx  # Same image for smoother motion
                    end_frame_reason = "single image mode, same frame for interpolation"
                else:
                    is_last_clip = (i == len(clip_info) - 1)
                    
                    # Check if we're in auto-cycle mode (no explicit scenes defined)
                    auto_cycle_mode = scenes_data is None or len(scenes_data) == 0
                    
                    # Track if scene transition already determined the end frame
                    scene_transition_handled = False
                    
                    if not is_last_clip:
                        next_info = clip_info[i + 1]
                        next_scene = next_info["scene_index"]
                        next_image_idx = next_info["image_idx"]
                        
                        if auto_cycle_mode:
                            # AUTO-CYCLE MODE: Check if next clip uses a different image
                            if next_image_idx != start_idx:
                                # Different image - blend to it
                                use_end_frame = True
                                actual_end_idx = next_image_idx
                                end_frame_reason = f"auto-cycle: blend to next image {next_image_idx + 1}"
                                scene_transition_handled = True
                        elif next_scene != scene_index:
                            # STORYBOARD MODE: Next clip is in DIFFERENT scene
                            next_transition = next_info["scene_transition"]
                            
                            # If transition is "blend" (or None), use next scene's image (scene transition priority)
                            if next_transition != "cut":
                                use_end_frame = True
                                actual_end_idx = next_info["image_idx"]
                                end_frame_reason = f"scene transition to scene {next_scene} (blend to next scene)"
                                scene_transition_handled = True
                            else:
                                # CUT transition: No end frame interpolation
                                use_end_frame = False
                                end_frame_reason = f"scene transition to scene {next_scene} (CUT - no interpolation)"
                                scene_transition_handled = True
                    
                    # Apply clip_mode logic if:
                    # - Scene transition didn't handle it (same scene, or different scene with "cut")
                    # - Or it's the last clip
                    if not scene_transition_handled:
                        if is_last_clip and last_frame_index is not None and last_frame_index < len(images):
                            # LAST CLIP with explicit end frame set
                            use_end_frame = True
                            actual_end_idx = last_frame_index
                            end_frame_reason = f"last clip with explicit end frame (image {last_frame_index + 1})"
                        elif is_last_clip and auto_cycle_mode:
                            # LAST CLIP in auto-cycle mode: cycle back to first available different image
                            # Find next different image (wrap around)
                            for offset in range(1, len(images)):
                                next_idx = (start_idx + offset) % len(images)
                                if next_idx != start_idx:
                                    use_end_frame = True
                                    actual_end_idx = next_idx
                                    end_frame_reason = f"last clip in auto-cycle: blend to image {next_idx + 1}"
                                    break
                            else:
                                # Only one image - no interpolation
                                use_end_frame = False
                                end_frame_reason = "last clip: single image, no interpolation"
                        elif is_last_clip:
                            # LAST CLIP in storyboard mode without explicit end frame
                            # NO end frame - clip ends naturally
                            use_end_frame = False
                            end_frame_reason = "last clip (storyboard mode), no end frame"
                        elif clip_mode == "blend":
                            # BLEND mode: Use next different image in cycle
                            # NOT same image - that causes same-frame generation issues
                            if auto_cycle_mode:
                                # Find next different image
                                for offset in range(1, len(images)):
                                    next_idx = (start_idx + offset) % len(images)
                                    if next_idx != start_idx:
                                        use_end_frame = True
                                        actual_end_idx = next_idx
                                        end_frame_reason = f"blend mode: cycle to image {next_idx + 1}"
                                        break
                                else:
                                    use_end_frame = False
                                    end_frame_reason = "blend mode: single image, no interpolation"
                            else:
                                # Storyboard blend mode - use same image for smooth motion within scene
                                use_end_frame = True
                                actual_end_idx = start_idx
                                end_frame_reason = "blend mode: same image for interpolation"
                        else:
                            # FRESH or CONTINUE mode (non-last clip): No end frame
                            use_end_frame = False
                            if clip_mode == "fresh":
                                end_frame_reason = "fresh mode, no end frame"
                            else:
                                end_frame_reason = "continue mode, no end frame"
                
                # Set frame names
                start_frame_name = images[actual_start_idx].name
                
                if use_end_frame and actual_end_idx is not None:
                    end_frame_name = images[actual_end_idx].name
                else:
                    end_frame_name = None
                    actual_end_idx = actual_start_idx  # For compatibility, but won't be used
                
                info["start_frame"] = start_frame_name
                info["end_frame"] = end_frame_name
                info["start_idx"] = actual_start_idx
                info["end_idx"] = actual_end_idx if use_end_frame else None
                info["use_end_frame"] = use_end_frame
                
                print(f"[Worker] Clip {i}: {start_frame_name} → {end_frame_name if end_frame_name else 'NONE'} (mode={clip_mode}, reason={end_frame_reason})", flush=True)
                
                # Determine initial status
                # For "continue" mode clips (except first in scene), set to WAITING_APPROVAL
                initial_status = ClipStatus.PENDING.value
                if info["requires_previous"]:
                    initial_status = ClipStatus.WAITING_APPROVAL.value
                    print(f"[Worker] Clip {i}: Set to WAITING_APPROVAL (requires previous clip approval)", flush=True)
                
                # Create clip record
                clip = Clip(
                    job_id=job_id,
                    clip_index=i,
                    dialogue_id=info["dialogue_id"],
                    dialogue_text=info["text"],
                    status=initial_status,
                    start_frame=start_frame_name,
                    end_frame=end_frame_name,
                )
                db.add(clip)
            
            db.commit()
        
        # Build clip_frames list for processing
        clip_frames = []
        # CRITICAL: Store original frame names BEFORE any modifications
        # These will be used for DB storage - NEVER overwritten with extracted frames
        original_clip_frames = {}
        
        for i, info in enumerate(clip_info):
            start_frame = images[info["start_idx"]]
            
            # Only set end_frame if this clip should use interpolation
            if info.get("use_end_frame") and info.get("end_idx") is not None:
                end_frame = images[info["end_idx"]]
            else:
                end_frame = None
            
            # Store ORIGINAL frame names (these NEVER change)
            original_clip_frames[i] = {
                "start_frame": start_frame.name if hasattr(start_frame, 'name') else str(start_frame),
                "end_frame": end_frame.name if end_frame and hasattr(end_frame, 'name') else None,
            }
            
            clip_frames.append({
                "start_index": info["start_idx"],
                "start_frame": start_frame,
                "end_index": info["end_idx"],
                "end_frame": end_frame,
                "clip_mode": info["clip_mode"],
                "requires_previous": info["requires_previous"],
                "scene_index": info["scene_index"],
                "original_scene_idx": info["image_idx"],  # Original scene image index for subject description
            })
        
        print(f"[Worker] Original clip frames preserved: {original_clip_frames}", flush=True)
        
        # VALIDATION: Prevent same-frame assignments (start == end)
        # EXCEPTION: Single image mode WITH interpolation - same frame is intentional
        for i, cf in enumerate(clip_frames):
            start_frame = cf["start_frame"]
            end_frame = cf["end_frame"]
            if end_frame is not None:
                # Check if start and end are the same
                same_frame = False
                if start_frame == end_frame:
                    same_frame = True
                elif hasattr(start_frame, 'name') and hasattr(end_frame, 'name') and start_frame.name == end_frame.name:
                    same_frame = True
                
                if same_frame:
                    # In single image mode with interpolation, same frame is intentional - keep it
                    if single_image_mode and use_interpolation:
                        print(f"[Worker] Clip {i}: Same start/end frame is OK (single image interpolation mode)", flush=True)
                    else:
                        print(f"[Worker] WARNING: Clip {i} has same start/end frame ({start_frame.name if hasattr(start_frame, 'name') else start_frame}), finding different end...", flush=True)
                        # Find a different end frame
                        start_idx = cf["start_index"]
                        for offset in range(1, len(images)):
                            next_idx = (start_idx + offset) % len(images)
                            next_img = images[next_idx]
                            if next_img != start_frame:
                                cf["end_frame"] = next_img
                                cf["end_index"] = next_idx
                                print(f"[Worker] Clip {i}: Changed end frame to {next_img.name}", flush=True)
                                break
                        else:
                            # No different frame available - set end_frame to None
                            cf["end_frame"] = None
                            print(f"[Worker] Clip {i}: No different frame available, setting end_frame to None", flush=True)
        
        # Log complete frame assignment summary
        print(f"\n{'='*70}", flush=True)
        print(f"[Worker] FRAME ASSIGNMENT SUMMARY", flush=True)
        print(f"{'='*70}", flush=True)
        print(f"Total clips: {len(clip_frames)}", flush=True)
        print(f"Last Frame Index: {last_frame_index}", flush=True)
        print(f"", flush=True)
        for i, cf in enumerate(clip_frames):
            mode = cf["clip_mode"]
            req_prev = cf["requires_previous"]
            start = cf["start_frame"].name if hasattr(cf["start_frame"], 'name') else str(cf["start_frame"])
            end = cf["end_frame"].name if cf["end_frame"] and hasattr(cf["end_frame"], 'name') else ("NONE" if cf["end_frame"] is None else str(cf["end_frame"]))
            status = "WAITING_APPROVAL" if req_prev else "PENDING"
            
            print(f"  Clip {i}: [{mode.upper()}] {start} → {end}", flush=True)
            print(f"           requires_previous={req_prev}, status={status}", flush=True)
            if mode == "continue":
                if req_prev:
                    print(f"           → Will extract start frame from clip {i-1} at runtime", flush=True)
                else:
                    print(f"           → First of scene, will use original image", flush=True)
        print(f"{'='*70}\n", flush=True)
        
        # Track completed AND APPROVED clips for 'continue' mode frame extraction
        approved_clip_videos = {}  # clip_index -> video_path (only approved ones)
        completed_clip_videos = {}  # clip_index -> video_path (all completed, for tracking)
        
        # Track subject descriptions per scene for continue mode consistency
        scene_subject_descriptions = {}  # scene_index -> subject description (generated on first clip)
        
        def get_or_generate_subject_description(scene_index: int, scene_image_path: Path) -> str:
            """Get cached subject description or generate for first clip of scene"""
            if scene_index in scene_subject_descriptions:
                return scene_subject_descriptions[scene_index]
            
            # Generate subject description from scene's original image
            print(f"[Worker] Generating subject description for scene {scene_index} from {scene_image_path.name}", flush=True)
            # Note: describe_subject_for_continuity is a stub that returns empty string
            # Keeping the call for future implementation
            description = describe_subject_for_continuity(str(scene_image_path))
            
            if description:
                scene_subject_descriptions[scene_index] = description
                print(f"[Worker] Scene {scene_index} subject: '{description}'", flush=True)
                
                # Log to database
                with get_db() as db:
                    add_job_log(
                        db, job_id,
                        f"📷 Scene {scene_index + 1} subject description: {description}",
                        "INFO", "prompt"
                    )
            else:
                scene_subject_descriptions[scene_index] = ""
                print(f"[Worker] Scene {scene_index}: No subject description generated", flush=True)
            
            return scene_subject_descriptions.get(scene_index, "")
        
        # Queue of pending clip indices (only PENDING status, not WAITING_APPROVAL)
        pending_clips = [i for i, info in enumerate(clip_info) if not info["requires_previous"]]
        waiting_clips = [i for i, info in enumerate(clip_info) if info["requires_previous"]]
        
        print(f"[Worker] Initial queue: {len(pending_clips)} pending, {len(waiting_clips)} waiting for approval", flush=True)
        
        def check_keys_available():
            """Check if any API keys are available using the KeyPoolManager"""
            from config import key_pool
            status = key_pool.get_pool_status_summary(generator.api_keys)
            return status["available"] > 0
        
        def send_no_keys_alert(job_id: str, retry_count: int):
            """Alert admin that keys are exhausted"""
            from config import key_pool
            status = key_pool.get_pool_status_summary(generator.api_keys)
            
            alert_msg = f"🚨 API KEYS EXHAUSTED - Job {job_id[:8]} paused (retry {retry_count}/{max_no_keys_retries})"
            print(f"\n{'='*60}", flush=True)
            print(alert_msg, flush=True)
            print(f"[KeyPool] Status: {status['available']} available, {status['rate_limited']} rate-limited, {status['invalid']} invalid", flush=True)
            print(f"{'='*60}\n", flush=True)
            
            # Log to database
            with get_db() as db:
                add_job_log(
                    db, job_id,
                    f"⚠️ All API keys exhausted! {status['rate_limited']} rate-limited, waiting {no_keys_wait_seconds}s (attempt {retry_count}/{max_no_keys_retries})",
                    "WARNING", "system",
                    details={"pool_status": status}
                )
            
            # Broadcast to UI
            self._broadcast_event(job_id, {
                "type": "keys_exhausted",
                "retry_count": retry_count,
                "max_retries": max_no_keys_retries,
                "wait_seconds": no_keys_wait_seconds,
                "message": f"All API keys exhausted ({status['rate_limited']} rate-limited). Waiting {no_keys_wait_seconds//60} min... ({retry_count}/{max_no_keys_retries})"
            })
            
            # Send email alert
            total_keys = status['total']
            send_key_alert_email("no_keys", 0, total_keys, job_id)
        
        def check_redo_clips():
            """Check for clips queued for redo and return their indices.
            NOTE: We only CHECK for redos here - the actual processing is handled by
            the independent _check_redo_queue() in the main worker loop, which starts
            redos immediately in separate threads.
            """
            redo_indices = []
            with get_db() as db:
                redo_clips = db.query(Clip).filter(
                    Clip.job_id == job_id,
                    Clip.status == ClipStatus.REDO_QUEUED.value
                ).all()
                
                for clip in redo_clips:
                    redo_indices.append(clip.clip_index)
                    # DON'T change status here - let the independent _check_redo_queue() handle it
                    # This allows redos to start immediately instead of waiting for the main loop
                
                # Don't log here - it spams when called repeatedly
                # The actual redo start is logged in _check_redo_queue()
            
            return redo_indices
        
        def extract_frame_from_video(video_path: Path, frame_offset: int = -8) -> Optional[Path]:
            """Extract a frame from video. frame_offset=-8 means 8 frames from the end."""
            try:
                # Use same ffmpeg/ffprobe config as video_processor.py
                # Also check ImageIO_FFMPEG_EXE as fallback (used in some setups)
                ffmpeg_exe = os.environ.get("FFMPEG_BIN") or os.environ.get("ImageIO_FFMPEG_EXE") or "ffmpeg"
                ffprobe_exe = os.environ.get("FFPROBE_BIN", "ffprobe")
                
                # If we have a custom ffmpeg path but not ffprobe, derive ffprobe from ffmpeg path
                if ffmpeg_exe not in ("ffmpeg", None) and ffprobe_exe == "ffprobe":
                    ffmpeg_path = Path(ffmpeg_exe)
                    if ffmpeg_path.exists():
                        # ffprobe should be in same directory
                        probe_name = "ffprobe.exe" if os.name == 'nt' else "ffprobe"
                        derived_probe = ffmpeg_path.parent / probe_name
                        if derived_probe.exists():
                            ffprobe_exe = str(derived_probe)
                
                print(f"[Worker] Using ffprobe: {ffprobe_exe}", flush=True)
                print(f"[Worker] Using ffmpeg: {ffmpeg_exe}", flush=True)
                
                # Get video duration
                probe_cmd = [
                    ffprobe_exe, "-v", "error", "-select_streams", "v:0",
                    "-show_entries", "stream=duration,r_frame_rate",
                    "-of", "csv=p=0", str(video_path)
                ]
                probe_result = subprocess.run(probe_cmd, capture_output=True, text=True)
                if probe_result.returncode != 0:
                    print(f"[Worker] ffprobe failed: {probe_result.stderr}", flush=True)
                    return None
                
                # Parse duration and fps
                parts = probe_result.stdout.strip().split(',')
                if len(parts) < 2:
                    print(f"[Worker] Could not parse ffprobe output: {probe_result.stdout}", flush=True)
                    return None
                    
                fps_str = parts[0]
                duration_str = parts[1] if len(parts) > 1 else "8"
                
                # Calculate fps from fraction (e.g., "30000/1001" or "30/1")
                if '/' in fps_str:
                    num, den = fps_str.split('/')
                    fps = float(num) / float(den)
                else:
                    fps = float(fps_str) if fps_str else 30.0
                
                duration = float(duration_str) if duration_str else 8.0
                
                # Calculate timestamp for frame_offset from end
                # frame_offset = -8 means 8 frames before end
                frames_from_end = abs(frame_offset)
                seconds_from_end = frames_from_end / fps
                timestamp = max(0, duration - seconds_from_end)
                
                print(f"[Worker] Extracting frame at {timestamp:.3f}s (fps={fps:.2f}, duration={duration:.2f}s, offset={frame_offset})", flush=True)
                
                # Extract frame
                output_frame = video_path.parent / f"{video_path.stem}_lastframe.jpg"
                extract_cmd = [
                    ffmpeg_exe, "-y", "-ss", str(timestamp), "-i", str(video_path),
                    "-frames:v", "1", "-q:v", "2", str(output_frame)
                ]
                extract_result = subprocess.run(extract_cmd, capture_output=True, text=True)
                
                if extract_result.returncode == 0 and output_frame.exists():
                    print(f"[Worker] Extracted frame to {output_frame.name}", flush=True)
                    return output_frame
                else:
                    print(f"[Worker] ffmpeg frame extraction failed: {extract_result.stderr}", flush=True)
                    return None
                    
            except Exception as e:
                print(f"[Worker] Frame extraction error: {e}", flush=True)
                import traceback
                traceback.print_exc()
                return None
        
        def enhance_frame_with_nano_banana(frame_path: Path, original_scene_image: Optional[Path] = None) -> Optional[Path]:
            """
            Enhance an extracted frame using Nano Banana Pro (Gemini 3 Pro Image).
            Upscales and improves quality of the image.
            
            If original_scene_image is provided, also corrects facial features to match
            the original person (fixes AI drift in facial appearance).
            """
            try:
                import google.genai as genai
                from google.genai import types
                import base64
                
                # Get API key — dedicated NANO_BANANA_API_KEY first, then fall back to general Gemini keys
                api_key = os.environ.get("NANO_BANANA_API_KEY", "").strip()
                if not api_key:
                    # Fallback to general Gemini keys pool
                    api_keys = get_gemini_keys_from_env()
                    if not api_keys:
                        print("[Worker] No API key available for Nano Banana Pro enhancement (set NANO_BANANA_API_KEY env var)", flush=True)
                        return frame_path  # Return original if no API key
                    api_key = api_keys[0]
                    print(f"[Worker] Using general Gemini key for Nano Banana (consider setting NANO_BANANA_API_KEY for a dedicated key)", flush=True)
                else:
                    print(f"[Worker] Using dedicated NANO_BANANA_API_KEY for enhancement", flush=True)
                
                client = genai.Client(api_key=api_key)
                
                # Read the extracted frame
                with open(frame_path, 'rb') as f:
                    frame_bytes = f.read()
                
                # Determine mime type
                suffix = frame_path.suffix.lower()
                mime_type = {
                    '.jpg': 'image/jpeg',
                    '.jpeg': 'image/jpeg',
                    '.png': 'image/png',
                    '.webp': 'image/webp'
                }.get(suffix, 'image/jpeg')
                
                print(f"[Worker] Enhancing frame with Nano Banana Pro: {frame_path.name}", flush=True)
                
                # Build the prompt parts
                parts = [
                    types.Part.from_bytes(data=frame_bytes, mime_type=mime_type),
                ]
                
                # If we have original scene image, include it for facial consistency
                if original_scene_image and original_scene_image.exists():
                    print(f"[Worker] Including original scene image for facial consistency: {original_scene_image.name}", flush=True)
                    
                    with open(original_scene_image, 'rb') as f:
                        original_bytes = f.read()
                    
                    original_suffix = original_scene_image.suffix.lower()
                    original_mime = {
                        '.jpg': 'image/jpeg',
                        '.jpeg': 'image/jpeg',
                        '.png': 'image/png',
                        '.webp': 'image/webp'
                    }.get(original_suffix, 'image/jpeg')
                    
                    parts.append(types.Part.from_bytes(data=original_bytes, mime_type=original_mime))
                    
                    prompt_text = (
                        "The first image is an extracted video frame. The second image shows the original person. "
                        "Enhance the first image while correcting the facial features to match the original person in the second image. "
                        "This is NOT a face swap - it's the same person, but the AI video generation may have slightly altered their appearance. "
                        "Correct any facial drift: restore accurate facial structure, skin tone, eye shape, nose shape, and other features to match the original. "
                        "Also upscale to higher resolution, reduce compression artifacts, and improve overall image quality. "
                        "Keep the exact pose, expression, lighting, background, and composition from the first image - only correct the facial features and enhance quality."
                    )
                else:
                    # No reference image - just enhance quality
                    prompt_text = (
                        "Upscale this image to higher resolution while preserving all details. "
                        "Enhance the image quality, reduce any compression artifacts, "
                        "and improve sharpness and clarity. Keep the exact same content, "
                        "colors, and composition - only improve the quality."
                    )
                
                parts.append(types.Part.from_text(text=prompt_text))
                
                contents = [
                    types.Content(
                        role="user",
                        parts=parts
                    )
                ]
                
                # Configure for image output
                config = types.GenerateContentConfig(
                    response_modalities=["IMAGE"],
                    temperature=0.2  # Low temperature for faithful reproduction
                )
                
                # Generate enhanced image with retry for 503 errors
                max_retries = 3
                response = None
                
                for attempt in range(max_retries):
                    try:
                        response = client.models.generate_content(
                            model="gemini-3-pro-image-preview",  # Nano Banana Pro
                            contents=contents,
                            config=config
                        )
                        break  # Success, exit retry loop
                    except Exception as api_error:
                        error_str = str(api_error)
                        if "503" in error_str or "overloaded" in error_str.lower():
                            if attempt < max_retries - 1:
                                wait_time = (attempt + 1) * 5  # 5s, 10s, 15s
                                print(f"[Worker] Nano Banana Pro overloaded, retrying in {wait_time}s (attempt {attempt + 1}/{max_retries})", flush=True)
                                time.sleep(wait_time)
                            else:
                                print(f"[Worker] Nano Banana Pro still overloaded after {max_retries} attempts, using original frame", flush=True)
                                return frame_path
                        else:
                            raise  # Re-raise non-503 errors
                
                if response is None:
                    print("[Worker] Nano Banana Pro: No response received, using original frame", flush=True)
                    return frame_path
                
                # Extract enhanced image from response
                enhanced_path = frame_path.parent / f"{frame_path.stem}_enhanced.png"
                
                if response.candidates and response.candidates[0].content.parts:
                    for part in response.candidates[0].content.parts:
                        if hasattr(part, 'inline_data') and part.inline_data:
                            # Save enhanced image
                            with open(enhanced_path, 'wb') as f:
                                f.write(part.inline_data.data)
                            print(f"[Worker] Enhanced frame saved to {enhanced_path.name}", flush=True)
                            return enhanced_path
                
                print("[Worker] Nano Banana Pro did not return an image, using original frame", flush=True)
                return frame_path
                
            except ImportError as e:
                print(f"[Worker] google-genai SDK not available for enhancement: {e}", flush=True)
                return frame_path
            except Exception as e:
                error_str = str(e)
                if "503" in error_str or "overloaded" in error_str.lower():
                    print(f"[Worker] Frame enhancement skipped (Nano Banana Pro overloaded), using original frame", flush=True)
                else:
                    print(f"[Worker] Frame enhancement error: {e}", flush=True)
                return frame_path  # Return original on error
        
        def process_single_clip(clip_index: int):
            """Process a single clip - runs in thread"""
            print(f"[Worker] process_single_clip({clip_index}) STARTED in thread", flush=True)
            
            if generator.cancelled:
                print(f"[Worker] Clip {clip_index}: Cancelled at start", flush=True)
                return {"clip_index": clip_index, "success": False, "skipped": True}
            
            # Stagger API calls to avoid hitting rate limits
            # Each clip waits 0-2 seconds based on its index
            import random
            stagger_delay = (clip_index % 3) * 0.5 + random.uniform(0, 0.5)
            print(f"[Worker] Clip {clip_index}: Stagger delay {stagger_delay:.2f}s", flush=True)
            time.sleep(stagger_delay)
            print(f"[Worker] Clip {clip_index}: Stagger delay complete", flush=True)
            
            # Check again after delay
            if generator.cancelled:
                print(f"[Worker] Clip {clip_index}: Cancelled after stagger", flush=True)
                return {"clip_index": clip_index, "success": False, "skipped": True}
            
            line_data = dialogue_data[clip_index]
            dialogue_id = line_data["id"]
            dialogue_text = line_data["text"]
            frames = clip_frames[clip_index]
            
            start_frame = frames["start_frame"]
            end_frame = frames["end_frame"]  # Can be None if no interpolation needed
            start_index = frames["start_index"]
            end_index = frames["end_index"]  # Can be None if no interpolation needed
            clip_mode = frames.get("clip_mode", "blend")
            requires_previous = frames.get("requires_previous", False)
            scene_index = frames.get("scene_index", 0)
            original_scene_idx = frames.get("original_scene_idx", 0)
            
            # For CONTINUE mode clips, inject subject description for visual consistency
            if clip_mode == "continue":
                # Get the original scene image (not extracted frame)
                scene_image_for_desc = images[original_scene_idx] if original_scene_idx < len(images) else images[0]
                subject_desc = get_or_generate_subject_description(scene_index, scene_image_for_desc)
                
                if subject_desc and requires_previous:
                    # Prepend subject description to dialogue for continue clips
                    # Format: "The [subject description] [dialogue]"
                    dialogue_text = f"{subject_desc} {dialogue_text}"
                    print(f"[Worker] Clip {clip_index}: Injected subject description for continuity", flush=True)
            
            # Store the original scene image for facial consistency correction
            original_scene_image = frames["start_frame"]  # The original image for this scene
            
            # Handle "continue" mode - use extracted frame from APPROVED previous clip
            # ONLY if requires_previous is True (meaning previous clip is in SAME scene)
            if clip_mode == "continue" and requires_previous and clip_index > 0:
                prev_idx = clip_index - 1
                prev_video = approved_clip_videos.get(prev_idx)
                print(f"[Worker] Clip {clip_index}: Continue mode check - prev_idx={prev_idx}, approved_clip_videos keys={list(approved_clip_videos.keys())}", flush=True)
                print(f"[Worker] Clip {clip_index}: prev_video={prev_video}", flush=True)
                if prev_video:
                    video_exists = Path(prev_video).exists()
                    print(f"[Worker] Clip {clip_index}: Video exists at path? {video_exists}", flush=True)
                    if video_exists:
                        extracted = extract_frame_from_video(Path(prev_video), frame_offset=-8)
                        if extracted:
                            # Enhance the extracted frame using Nano Banana Pro
                            # Pass the original scene image for facial consistency correction
                            enhanced = enhance_frame_with_nano_banana(extracted, original_scene_image)
                            start_frame = enhanced
                            print(f"[Worker] Clip {clip_index}: Using {'enhanced' if enhanced != extracted else 'extracted'} frame from APPROVED clip {prev_idx}", flush=True)
                        else:
                            print(f"[Worker] Clip {clip_index}: Frame extraction failed, using original image", flush=True)
                    else:
                        print(f"[Worker] Clip {clip_index}: Video file does not exist at {prev_video}, using original image", flush=True)
                else:
                    print(f"[Worker] Clip {clip_index}: Approved previous clip video not found (prev_video is None), using original image", flush=True)
            elif clip_mode == "continue" and not requires_previous:
                # First clip of scene in Continue mode - use original image
                print(f"[Worker] Clip {clip_index}: Continue mode but first of scene, using original image", flush=True)
            
            # Get frame names for logging/database (handle both Path objects and strings)
            def get_frame_name(frame):
                if frame is None:
                    return None
                if hasattr(frame, 'name'):
                    return frame.name
                if hasattr(frame, 'stem'):
                    return Path(frame).name
                return str(frame).split('/')[-1] if '/' in str(frame) else str(frame)
            
            # For generation, we use start_frame (which may be extracted frame for CONTINUE mode)
            # But for DATABASE STORAGE, we ALWAYS store the ORIGINAL scene image names
            # This ensures redo can find the correct images
            original_start_name = get_frame_name(original_scene_image)  # Always the original uploaded image
            original_end_name = get_frame_name(frames.get("end_frame")) if frames.get("end_frame") else None
            
            # For logging, show what we're actually using for generation
            actual_start_name = get_frame_name(start_frame)
            actual_end_name = get_frame_name(end_frame) if end_frame else None
            
            # Update clip status to generating
            print(f"[Worker] Clip {clip_index}: Updating status to GENERATING", flush=True)
            print(f"[Worker] Clip {clip_index}: DB will store: start='{original_start_name}', end='{original_end_name}'", flush=True)
            if actual_start_name != original_start_name:
                print(f"[Worker] Clip {clip_index}: Generation will use extracted frame: '{actual_start_name}'", flush=True)
            
            with get_db() as db:
                clip = db.query(Clip).filter(
                    Clip.job_id == job_id,
                    Clip.clip_index == clip_index
                ).first()
                
                if clip:
                    clip.status = ClipStatus.GENERATING.value
                    clip.started_at = datetime.utcnow()
                    # CRITICAL: Store ORIGINAL image names, not extracted frame names!
                    clip.start_frame = original_start_name
                    clip.end_frame = original_end_name
                    db.commit()
                    print(f"[Worker] Clip {clip_index}: Status updated to GENERATING", flush=True)
                else:
                    print(f"[Worker] Clip {clip_index}: WARNING - Clip record not found!", flush=True)
            
            # Log exact frame assignment for debugging
            print(f"[Worker] CLIP {clip_index} FRAME ASSIGNMENT:", flush=True)
            print(f"  - start_frame (for generation): {actual_start_name} (mode={clip_mode})", flush=True)
            print(f"  - end_frame (for generation): {actual_end_name if actual_end_name else 'NONE (no interpolation)'}", flush=True)
            print(f"  - original_start (stored in DB): {original_start_name}", flush=True)
            print(f"  - original_end (stored in DB): {original_end_name}", flush=True)
            
            self._broadcast_event(job_id, {
                "type": "clip_started",
                "clip_index": clip_index,
                "dialogue_id": dialogue_id,
                "start_frame": original_start_name,  # UI shows original, not extracted
                "end_frame": original_end_name,
            })
            
            # Check if start frame is blacklisted (only for Path objects, not extracted frames)
            if hasattr(start_frame, 'exists') and start_frame in generator.blacklist:
                result = self._get_next_clean_start(generator, images, start_index)
                if result:
                    start_index, start_frame = result
                else:
                    with get_db() as db:
                        clip = db.query(Clip).filter(
                            Clip.job_id == job_id,
                            Clip.clip_index == clip_index
                        ).first()
                        if clip:
                            clip.status = ClipStatus.FAILED.value
                            clip.error_code = "ALL_IMAGES_BLACKLISTED"
                            clip.error_message = "No clean images available"
                            db.commit()
                    return {"clip_index": clip_index, "success": False, "failed": True}
            
            # Generate clip
            try:
                print(f"[Worker] Generating clip {clip_index + 1}/{total_clips} in parallel...", flush=True)
                
                # Get the original scene image for prompt analysis
                # This is the uploaded image for this scene (not the extracted frame in CONTINUE mode)
                scene_image = images[start_index] if start_index < len(images) else images[0]
                
                # Calculate dynamic duration for LAST CLIP
                # Last clip picks from 4, 6, or 8 seconds based on expected speech duration
                override_duration = None
                is_last_clip = clip_index == total_clips - 1
                
                if is_last_clip:
                    # Estimate speech duration based on word count
                    word_count = len(dialogue_text.split())
                    language = generator.config.language if hasattr(generator.config, 'language') else 'English'
                    
                    # Words per second by language (approximate)
                    wps_map = {
                        "English": 2.5, "Italian": 2.8, "Spanish": 2.8, "French": 2.5, "German": 2.2,
                        "Portuguese": 2.7, "Dutch": 2.4, "Polish": 2.3, "Russian": 2.2,
                        "Japanese": 3.0, "Korean": 3.0, "Chinese": 3.2, "Arabic": 2.3, "Hindi": 2.6, "Turkish": 2.5
                    }
                    wps = wps_map.get(language, 2.5)
                    estimated_duration = word_count / wps
                    
                    # Pick the duration slightly above the estimated (4, 6, or 8 seconds)
                    if estimated_duration <= 3.5:
                        override_duration = "4"
                    elif estimated_duration <= 5.5:
                        override_duration = "6"
                    else:
                        override_duration = "8"
                    
                    print(f"[Worker] LAST CLIP: {word_count} words, ~{estimated_duration:.1f}s speech → using {override_duration}s duration", flush=True)
                
                # CRITICAL: Log the actual start_frame being used for generation
                actual_start_frame_name = start_frame.name if hasattr(start_frame, 'name') else str(start_frame)
                print(f"[Worker] >>> GENERATING with start_frame: {actual_start_frame_name}", flush=True)
                if clip_mode == "continue" and requires_previous:
                    print(f"[Worker] >>> (This should be an EXTRACTED frame from previous clip, NOT the scene image)", flush=True)
                
                result = generator.generate_single_clip(
                    start_frame=start_frame,
                    end_frame=end_frame,  # Can be None
                    dialogue_line=dialogue_text,
                    dialogue_id=dialogue_id,
                    clip_index=clip_index,
                    output_dir=output_dir,
                    images_list=images,
                    current_end_index=end_index if end_index is not None else start_index,
                    scene_image=scene_image,  # Original scene image for prompt analysis
                    override_duration=override_duration,  # Dynamic duration for last clip
                    generation_mode=generation_mode,  # Pass generation mode for blacklist scoping
                )
                
                # Log the FULL prompt that was sent to Veo (no truncation)
                if result.get("prompt_text"):
                    full_prompt = result["prompt_text"]
                    with get_db() as db:
                        add_job_log(
                            db, job_id,
                            f"📝 FULL PROMPT for clip {clip_index + 1}:\n{full_prompt}",
                            "INFO", "prompt"
                        )
                
                # Check if failed due to no keys or rate limit exhaustion
                if not result["success"]:
                    error = result.get("error")
                    if error and hasattr(error, 'code'):
                        if error.code.value in ["API_KEY_INVALID", "API_QUOTA_EXCEEDED", "RATE_LIMIT_429"]:
                            return {"clip_index": clip_index, "success": False, "no_keys": True, "result": result}
                        # Check for should_pause flag (keys exhausted after retries)
                        if hasattr(error, 'details') and error.details.get("should_pause"):
                            return {"clip_index": clip_index, "success": False, "no_keys": True, "should_pause": True, "result": result}
                
            except Exception as gen_error:
                print(f"[Worker] Clip {clip_index} CRASHED: {type(gen_error).__name__}: {str(gen_error)[:200]}", flush=True)
                result = {
                    "success": False,
                    "error": gen_error,
                    "output_path": None,
                    "end_frame_used": None,
                    "end_index": end_index,
                }
            
            # Update clip record
            try:
                with get_db() as db:
                    clip = db.query(Clip).filter(
                        Clip.job_id == job_id,
                        Clip.clip_index == clip_index
                    ).first()
                    
                    if clip:
                        clip.completed_at = datetime.utcnow()
                        
                        if clip.started_at:
                            clip.duration_seconds = (
                                clip.completed_at - clip.started_at
                            ).total_seconds()
                        
                        if result["success"]:
                            new_filename = result["output_path"].name if result.get("output_path") else None
                            
                            versions = [{
                                "attempt": 1,
                                "filename": new_filename,
                                "generated_at": datetime.utcnow().isoformat(),
                            }]
                            clip.versions_json = json.dumps(versions)
                            clip.selected_variant = 1
                            
                            clip.status = ClipStatus.COMPLETED.value
                            clip.approval_status = "pending_review"
                            clip.output_filename = new_filename
                            clip.prompt_text = result.get("prompt_text")
                            
                            # Track video path for reference (NOT in approved_clip_videos yet - user must approve first)
                            if result.get("output_path"):
                                video_path = str(result["output_path"])
                                completed_clip_videos[clip_index] = video_path
                                # NOTE: Don't add to approved_clip_videos here!
                                # That happens when user approves (in waiting_clips check)
                                
                                # Upload to R2 for persistence (API jobs)
                                # This ensures videos survive server restarts on ephemeral platforms
                                try:
                                    from backends.storage import is_storage_configured, get_storage
                                    if is_storage_configured():
                                        storage = get_storage()
                                        r2_key = f"jobs/{job_id}/outputs/{new_filename}"
                                        storage.upload_file(video_path, r2_key, content_type='video/mp4')
                                        # Get presigned URL for UI access
                                        output_url = storage.get_presigned_url(r2_key, expires_in=86400 * 7)
                                        clip.output_url = output_url
                                        # Update version entry with URL
                                        versions[0]["url"] = output_url
                                        clip.versions_json = json.dumps(versions)
                                        print(f"[Worker] Uploaded clip {clip_index} to R2: {r2_key}", flush=True)
                                except Exception as r2_err:
                                    print(f"[Worker] R2 upload failed for clip {clip_index} (non-fatal): {r2_err}", flush=True)
                                    # Non-fatal - local file still exists
                        elif result.get("skipped") and result.get("skip_reason") == "celebrity_filter":
                            # Celebrity filter triggered - mark as skipped
                            clip.status = ClipStatus.SKIPPED.value
                            clip.error_code = "CELEBRITY_FILTER"
                            clip.error_message = "Skipped due to celebrity filter - eligible for reimbursement"
                            clip.prompt_text = result.get("prompt_text") or result.get("result", {}).get("prompt_text")
                            
                            # Log for user
                            add_job_log(
                                db, job_id,
                                f"⚠️ Clip {clip_index + 1} skipped (celebrity filter). Eligible for reimbursement.",
                                "WARNING", "celebrity_filter"
                            )
                        else:
                            # Check if this is a "no keys" situation - re-queue as redo
                            if result.get("no_keys") or result.get("should_pause"):
                                clip.status = ClipStatus.REDO_QUEUED.value
                                add_job_log(
                                    db, job_id,
                                    f"Clip {clip_index + 1} re-queued: API keys temporarily unavailable",
                                    "WARNING", "system"
                                )
                            else:
                                clip.status = ClipStatus.FAILED.value
                                error_obj = result.get("error")
                                if error_obj:
                                    clip.error_code = error_obj.code.value if hasattr(error_obj, 'code') else "UNKNOWN"
                                    clip.error_message = str(error_obj.message if hasattr(error_obj, 'message') else error_obj)[:500]
                        
                        db.commit()
                    
                    # Save generation log if successful
                    if result.get("success") and result.get("output_path"):
                        # Safely get frame names (handle None cases for single-image mode)
                        start_frame_name = start_frame.name if start_frame and hasattr(start_frame, 'name') else str(start_frame) if start_frame else "unknown"
                        
                        # For end frame: prefer result's end_frame_used, then end_frame, or fall back to start_frame
                        if result.get("end_frame_used") and hasattr(result["end_frame_used"], 'name'):
                            end_frame_name = result["end_frame_used"].name
                        elif end_frame and hasattr(end_frame, 'name'):
                            end_frame_name = end_frame.name
                        else:
                            end_frame_name = start_frame_name  # Single image mode fallback
                        
                        gen_log = GenerationLog(
                            job_id=job_id,
                            video_id=dialogue_id,
                            images_dir=images_dir_str,
                            start_frame=start_frame_name,
                            end_frame=end_frame_name,
                            dialogue_line=dialogue_text,
                            language=generator.config.language,
                            prompt_text=result.get("prompt_text", ""),
                            video_filename=result["output_path"].name,
                            aspect_ratio=generator.config.aspect_ratio if isinstance(generator.config.aspect_ratio, str) else generator.config.aspect_ratio.value,
                            resolution=generator.config.resolution if isinstance(generator.config.resolution, str) else generator.config.resolution.value,
                            duration=generator.config.duration if isinstance(generator.config.duration, str) else generator.config.duration.value,
                        )
                        db.add(gen_log)
                        db.commit()
            except Exception as db_error:
                print(f"[Worker] DB error updating clip {clip_index}: {db_error}")
            
            self._broadcast_event(job_id, {
                "type": "clip_completed" if result["success"] else ("clip_skipped" if result.get("skipped") else "clip_failed"),
                "clip_index": clip_index,
                "success": result["success"],
                "skipped": result.get("skipped", False),
                "output": result["output_path"].name if result.get("output_path") else None,
            })
            
            return {
                "clip_index": clip_index,
                "success": result["success"],
                "skipped": result.get("skipped", False),
                "skip_reason": result.get("skip_reason"),
                "result": result,
            }
        
        # Get generation mode from config
        generation_mode = getattr(generator.config, 'generation_mode', 'parallel')
        
        # OVERRIDE: Force parallel mode when:
        # - Single image (no frame chaining needed)
        # - Storyboard mode (has its own frame logic via scenes_data)
        if generation_mode != 'parallel':
            if len(images) == 1:
                generation_mode = 'parallel'
                print(f"[Worker] Single image detected - using parallel mode", flush=True)
            elif scenes_data and len(scenes_data) > 0:
                generation_mode = 'parallel'
                print(f"[Worker] Storyboard mode detected - using parallel mode", flush=True)
        
        print(f"[Worker] Generation mode: {generation_mode}", flush=True)
        
        # Log the generation mode
        with get_db() as db:
            if generation_mode == 'sequential':
                mode_emoji = "🔗"
                mode_desc = "sequential (one at a time)"
            elif generation_mode == 'staggered':
                mode_emoji = "🔀"
                mode_desc = "staggered parallel (chained frames, parallel generation)"
            else:
                mode_emoji = "⚡"
                mode_desc = "parallel (all at once)"
            add_job_log(
                db, job_id,
                f"{mode_emoji} Generation mode: {mode_desc}",
                "INFO", "config"
            )
        
        # === SEQUENTIAL MODE: Process clips one-by-one with frame chaining ===
        if generation_mode == 'sequential':
            print(f"[Worker] 🔗 SEQUENTIAL MODE: Processing clips one-by-one for guaranteed smooth transitions", flush=True)
            
            with get_db() as db:
                add_job_log(
                    db, job_id,
                    "🔗 Sequential mode: Clips will process one at a time for guaranteed smooth transitions",
                    "INFO", "system"
                )
            
            # Track current start frame (initially from clip_frames[0])
            current_start_frame = clip_frames[0]["start_frame"] if clip_frames else None
            current_start_index = clip_frames[0]["start_index"] if clip_frames else 0
            
            # Process all clips sequentially
            all_clip_indices = list(range(len(clip_frames)))
            processed_indices = set()
            
            while all_clip_indices and not generator.cancelled:
                # Check for redo clips first
                redo_indices = check_redo_clips()
                if redo_indices:
                    for idx in redo_indices:
                        if idx not in all_clip_indices and idx not in processed_indices:
                            # Insert at appropriate position (after current)
                            all_clip_indices.append(idx)
                            print(f"[Worker] Added redo clip {idx} to sequential queue", flush=True)
                
                # Check if keys are available
                if not check_keys_available():
                    if no_keys_retries == 0:
                        print(f"[Worker] ⚠️ NO KEYS AVAILABLE (sequential mode) - will pause", flush=True)
                    no_keys_retries += 1
                    
                    if no_keys_retries > max_no_keys_retries:
                        # Pause job instead of failing
                        with get_db() as db:
                            job = db.query(Job).filter(Job.id == job_id).first()
                            if job:
                                job.status = JobStatus.PAUSED.value
                                db.commit()
                            add_job_log(
                                db, job_id, 
                                f"⏸️ Job paused: API keys exhausted. Will auto-resume when keys available.",
                                "WARNING", "system"
                            )
                        generator.paused = True
                        raise JobPausedException("API keys exhausted - job paused")
                    
                    # Wait for keys
                    send_no_keys_alert(job_id, no_keys_retries)
                    wait_end = time.time() + no_keys_wait_seconds
                    while time.time() < wait_end and not generator.cancelled:
                        if check_keys_available():
                            print(f"[Worker] ✅ Keys available again, resuming sequential processing...", flush=True)
                            break
                        time.sleep(10)
                    continue
                
                no_keys_retries = 0
                
                # Get next clip to process
                clip_index = all_clip_indices.pop(0)
                processed_indices.add(clip_index)
                
                # Update the start frame for this clip based on previous results
                if clip_index > 0 and current_start_frame is not None:
                    # Use the actual end frame from previous clip as this clip's start
                    clip_frames[clip_index]["start_frame"] = current_start_frame
                    clip_frames[clip_index]["start_index"] = current_start_index
                    print(f"[Worker] Clip {clip_index}: Using chained start frame from previous clip: {current_start_frame.name if hasattr(current_start_frame, 'name') else current_start_frame}", flush=True)
                    
                    # CRITICAL: Also update end frame to be DIFFERENT from new start frame
                    # Find the next clean image that's not the start frame
                    current_end = clip_frames[clip_index].get("end_frame")
                    current_end_index = clip_frames[clip_index].get("end_index", current_start_index)
                    
                    # Check if end frame needs updating:
                    # 1. End frame is same as new start
                    # 2. End frame is in blacklist
                    # 3. End frame is None
                    needs_new_end = False
                    if current_end is None:
                        needs_new_end = True
                        reason = "end frame is None"
                    elif current_end == current_start_frame:
                        needs_new_end = True
                        reason = "end frame same as start (object match)"
                    elif hasattr(current_end, 'name') and hasattr(current_start_frame, 'name') and current_end.name == current_start_frame.name:
                        needs_new_end = True
                        reason = "end frame same as start (name match)"
                    elif current_end in generator.blacklist:
                        needs_new_end = True
                        reason = f"end frame {current_end.name if hasattr(current_end, 'name') else current_end} is blacklisted"
                    
                    if needs_new_end:
                        print(f"[Worker] Clip {clip_index}: {reason}, finding different end frame...", flush=True)
                        # Find next available image after current start
                        found_end = False
                        for offset in range(1, len(images)):
                            next_idx = (current_start_index + offset) % len(images)
                            next_img = images[next_idx]
                            if next_img != current_start_frame and next_img not in generator.blacklist:
                                clip_frames[clip_index]["end_frame"] = next_img
                                clip_frames[clip_index]["end_index"] = next_idx
                                print(f"[Worker] Clip {clip_index}: Updated end frame to {next_img.name}", flush=True)
                                found_end = True
                                break
                        if not found_end:
                            # No different clean frame found - log available images
                            available = [img.name for img in images if img not in generator.blacklist and img != current_start_frame]
                            print(f"[Worker] Clip {clip_index}: WARNING - Could not find different end frame", flush=True)
                            print(f"[Worker] Clip {clip_index}: Available images: {available}", flush=True)
                            print(f"[Worker] Clip {clip_index}: Blacklisted: {[img.name for img in generator.blacklist]}", flush=True)
                    
                    # NOTE: Do NOT update clip.start_frame/end_frame here!
                    # Clips are created with original frame names and those should be preserved.
                    # The current_start_frame may be an extracted frame for CONTINUE mode,
                    # which is correct for generation but should NOT be stored in DB.
                
                # Process this clip synchronously
                print(f"[Worker] Sequential: Processing clip {clip_index + 1}/{len(clip_frames)}", flush=True)
                result = process_single_clip(clip_index)
                
                if result.get("success"):
                    completed += 1
                    
                    # Get the actual end frame used and chain it to next clip
                    inner_result = result.get("result", {})
                    end_frame_used = inner_result.get("end_frame_used")
                    
                    if end_frame_used:
                        current_start_frame = end_frame_used
                        current_start_index = inner_result.get("end_index", current_start_index)
                        print(f"[Worker] Clip {clip_index}: Chaining end frame '{end_frame_used.name if hasattr(end_frame_used, 'name') else end_frame_used}' to next clip", flush=True)
                        
                        # Track completion (NOT approved yet - user must approve first)
                        video_path = str(inner_result.get("output_path")) if inner_result.get("output_path") else None
                        completed_clip_videos[clip_index] = video_path
                        # NOTE: Don't add to approved_clip_videos - that happens on user approval
                    else:
                        # No end frame - try to find next clean image for continuity
                        print(f"[Worker] Clip {clip_index}: No end frame returned, finding next clean image", flush=True)
                        next_result = self._get_next_clean_start(generator, images, current_start_index)
                        if next_result:
                            current_start_index, current_start_frame = next_result
                elif result.get("skipped"):
                    skipped += 1
                    completed_clip_videos[clip_index] = None  # Mark as done for dependents
                    print(f"[Worker] Sequential: Clip {clip_index} skipped, marking as done", flush=True)
                else:
                    failed += 1
                    # On failure, try to continue with next clean image
                    print(f"[Worker] Clip {clip_index} failed, finding next clean frame for continuity", flush=True)
                    next_result = self._get_next_clean_start(generator, images, current_start_index)
                    if next_result:
                        current_start_index, current_start_frame = next_result
                    completed_clip_videos[clip_index] = None
                
                # Update job progress
                with get_db() as db:
                    job = db.query(Job).filter(Job.id == job_id).first()
                    if job:
                        job.completed_clips = completed
                        job.failed_clips = failed
                        job.skipped_clips = skipped
                        processed = completed + failed + skipped
                        job.progress_percent = (processed / total_clips) * 100 if total_clips > 0 else 0
                        db.commit()
                
                # Small delay between clips to avoid rate limits
                time.sleep(1)
        
        # === STAGGERED MODE: Odd-Even processing for optimal speed with guaranteed transitions ===
        elif generation_mode == 'staggered':
            print(f"[Worker] 🔀 STAGGERED MODE: Odd-Even processing (odd clips first, then even)", flush=True)
            
            with get_db() as db:
                add_job_log(
                    db, job_id,
                    "🔀 Staggered mode: Processing odd clips first, then even clips with confirmed frames",
                    "INFO", "system"
                )
            
            from concurrent.futures import ThreadPoolExecutor, as_completed
            
            # Split clips into odd (0, 2, 4...) and even (1, 3, 5...) indices
            # Note: We use 0-indexed, so "odd indices" are actually clips 1, 3, 5...
            odd_indices = [i for i in range(0, len(clip_frames), 2)]  # 0, 2, 4...
            even_indices = [i for i in range(1, len(clip_frames), 2)]  # 1, 3, 5...
            
            print(f"[Worker] Phase 1: Odd clips {odd_indices}", flush=True)
            print(f"[Worker] Phase 2: Even clips {even_indices}", flush=True)
            
            # Track confirmed frames from each clip
            confirmed_frames = {}  # clip_index -> (start_frame, end_frame)
            
            def process_clip_for_staggered(clip_index: int, frames_locked: bool = False):
                """Process a single clip and return confirmed frames
                
                Args:
                    clip_index: The index of the clip to process
                    frames_locked: If True, frames cannot be swapped (Phase 2 - frames confirmed from Phase 1)
                """
                if generator.cancelled:
                    return {"clip_index": clip_index, "success": False, "skipped": True, "confirmed": None}
                
                frames = clip_frames[clip_index]
                line_data = dialogue_data[clip_index]
                
                start_frame = frames["start_frame"]
                end_frame = frames["end_frame"]
                start_index = frames["start_index"]
                end_index = frames["end_index"]
                dialogue_text = line_data["text"]
                dialogue_id = line_data["id"]
                
                # Debug: Log what frames we're actually using
                start_name = start_frame.name if hasattr(start_frame, 'name') else str(start_frame)
                end_name = end_frame.name if end_frame and hasattr(end_frame, 'name') else str(end_frame)
                lock_status = "LOCKED" if frames_locked else "unlocked"
                print(f"[Worker] process_clip_for_staggered({clip_index}, {lock_status}): Using frames {start_name} → {end_name}", flush=True)
                
                # Update clip status
                with get_db() as db:
                    clip = db.query(Clip).filter(
                        Clip.job_id == job_id,
                        Clip.clip_index == clip_index
                    ).first()
                    if clip:
                        clip.status = ClipStatus.GENERATING.value
                        clip.started_at = datetime.utcnow()
                        # NOTE: Do NOT update start_frame/end_frame here!
                        # They were set correctly at clip creation and should be preserved.
                        # The start_frame/end_frame variables may be modified for CONTINUE mode.
                        db.commit()
                
                # Broadcast with original frame names (for UI display)
                orig_frames = original_clip_frames.get(clip_index, {})
                self._broadcast_event(job_id, {
                    "type": "clip_started",
                    "clip_index": clip_index,
                    "start_frame": orig_frames.get("start_frame", ""),
                    "end_frame": orig_frames.get("end_frame"),
                })
                
                # Track the confirmed frames via callback
                confirmed_start = None
                confirmed_end = None
                
                def on_confirmed(idx, s, e):
                    nonlocal confirmed_start, confirmed_end
                    confirmed_start = s
                    confirmed_end = e
                
                try:
                    result = generator.generate_single_clip(
                        start_frame=start_frame,
                        end_frame=end_frame,
                        dialogue_line=dialogue_text,
                        dialogue_id=dialogue_id,
                        clip_index=clip_index,
                        output_dir=output_dir,
                        images_list=images,
                        current_end_index=end_index if end_index is not None else start_index,
                        generation_mode="staggered",
                        on_frames_locked=on_confirmed,
                        frames_locked=frames_locked,  # Phase 2 clips have locked frames
                    )
                except Exception as e:
                    print(f"[Worker] Clip {clip_index} generation error: {e}", flush=True)
                    result = {"success": False, "error": str(e)}
                
                # Update clip record
                with get_db() as db:
                    clip = db.query(Clip).filter(
                        Clip.job_id == job_id,
                        Clip.clip_index == clip_index
                    ).first()
                    if clip:
                        clip.completed_at = datetime.utcnow()
                        if result.get("success"):
                            new_filename = result["output_path"].name if result.get("output_path") else None
                            
                            # Set versions_json for first generation
                            versions = [{
                                "attempt": 1,
                                "filename": new_filename,
                                "generated_at": datetime.utcnow().isoformat(),
                            }]
                            clip.versions_json = json.dumps(versions)
                            clip.selected_variant = 1
                            
                            clip.status = ClipStatus.COMPLETED.value
                            clip.output_filename = new_filename
                            clip.approval_status = "pending_review"
                            
                            # Upload to R2 for persistence (API jobs)
                            if result.get("output_path"):
                                try:
                                    from backends.storage import is_storage_configured, get_storage
                                    if is_storage_configured():
                                        storage = get_storage()
                                        r2_key = f"jobs/{job_id}/outputs/{new_filename}"
                                        storage.upload_file(str(result["output_path"]), r2_key, content_type='video/mp4')
                                        output_url = storage.get_presigned_url(r2_key, expires_in=86400 * 7)
                                        clip.output_url = output_url
                                        versions[0]["url"] = output_url
                                        clip.versions_json = json.dumps(versions)
                                        print(f"[Worker] Uploaded clip {clip_index} to R2: {r2_key}", flush=True)
                                except Exception as r2_err:
                                    print(f"[Worker] R2 upload failed for clip {clip_index} (non-fatal): {r2_err}", flush=True)
                        else:
                            # Check if this is a "no keys" situation - re-queue as redo
                            if result.get("no_keys") or result.get("should_pause"):
                                clip.status = ClipStatus.REDO_QUEUED.value
                                add_job_log(
                                    db, job_id,
                                    f"Clip {clip_index + 1} re-queued: API keys temporarily unavailable",
                                    "WARNING", "system"
                                )
                            else:
                                clip.status = ClipStatus.FAILED.value
                                error_obj = result.get("error")
                                if error_obj:
                                    clip.error_message = str(error_obj)[:500]
                        db.commit()
                
                self._broadcast_event(job_id, {
                    "type": "clip_completed",
                    "clip_index": clip_index,
                    "success": result.get("success", False),
                    "output": result["output_path"].name if result.get("output_path") else None,
                })
                
                return {
                    "clip_index": clip_index,
                    "success": result.get("success", False),
                    "result": result,
                    "confirmed": (confirmed_start, confirmed_end) if confirmed_start else None,
                }
            
            # === PHASE 1: Process odd clips (0, 2, 4...) in parallel ===
            print(f"[Worker] === PHASE 1: Processing {len(odd_indices)} odd clips in parallel ===", flush=True)
            
            with get_db() as db:
                add_job_log(db, job_id, f"Phase 1: Processing odd clips {[i+1 for i in odd_indices]} in parallel", "INFO", "system")
            
            with ThreadPoolExecutor(max_workers=min(parallel_clips, len(odd_indices))) as executor:
                # Phase 1: frames_locked=False - frames can be swapped if celebrity filter triggers
                futures = {executor.submit(process_clip_for_staggered, i, False): i for i in odd_indices}
                
                for future in as_completed(futures):
                    clip_index = futures[future]
                    try:
                        result = future.result()
                        if result.get("success"):
                            completed += 1
                            if result.get("result", {}).get("output_path"):
                                completed_clip_videos[clip_index] = str(result["result"]["output_path"])
                        elif result.get("skipped"):
                            skipped += 1
                            completed_clip_videos[clip_index] = None  # Mark as done for dependents
                            print(f"[Worker] Phase 1: Clip {clip_index} skipped, marking as done", flush=True)
                        else:
                            failed += 1
                        
                        # Store confirmed frames
                        if result.get("confirmed"):
                            confirmed_frames[clip_index] = result["confirmed"]
                            start_name = result['confirmed'][0].name if result['confirmed'][0] and hasattr(result['confirmed'][0], 'name') else str(result['confirmed'][0])
                            end_name = result['confirmed'][1].name if result['confirmed'][1] and hasattr(result['confirmed'][1], 'name') else str(result['confirmed'][1])
                            print(f"[Worker] Clip {clip_index} confirmed: {start_name} → {end_name}", flush=True)
                            with get_db() as db:
                                add_job_log(db, job_id, f"Clip {clip_index+1} frames locked: {start_name} → {end_name}", "DEBUG", "system")
                        else:
                            print(f"[Worker] WARNING: Clip {clip_index} has NO confirmed frames!", flush=True)
                        
                        # Update progress
                        with get_db() as db:
                            job = db.query(Job).filter(Job.id == job_id).first()
                            if job:
                                job.completed_clips = completed
                                job.failed_clips = failed
                                job.skipped_clips = skipped
                                processed = completed + failed + skipped
                                job.progress_percent = (processed / total_clips) * 100 if total_clips > 0 else 0
                                db.commit()
                    except Exception as e:
                        print(f"[Worker] Phase 1 future error for clip {clip_index}: {e}", flush=True)
                        failed += 1
            
            print(f"[Worker] === PHASE 1 COMPLETE: {len(confirmed_frames)} clips confirmed ===", flush=True)
            
            # Debug: Print all confirmed frames
            for idx, frames in confirmed_frames.items():
                start_name = frames[0].name if frames[0] and hasattr(frames[0], 'name') else str(frames[0])
                end_name = frames[1].name if frames[1] and hasattr(frames[1], 'name') else str(frames[1])
                print(f"[Worker] DEBUG confirmed_frames[{idx}] = ({start_name} → {end_name})", flush=True)
            
            # === PHASE 2: Process even clips (1, 3, 5...) using confirmed frames ===
            if even_indices and not generator.cancelled:
                print(f"[Worker] === PHASE 2: Processing {len(even_indices)} even clips with confirmed frames ===", flush=True)
                
                # Update even clips' frames based on confirmed odd clips
                for clip_index in even_indices:
                    prev_idx = clip_index - 1  # Previous odd clip
                    next_idx = clip_index + 1  # Next odd clip
                    
                    print(f"[Worker] DEBUG Clip {clip_index}: prev_idx={prev_idx}, next_idx={next_idx}", flush=True)
                    print(f"[Worker] DEBUG Clip {clip_index}: prev_idx in confirmed_frames = {prev_idx in confirmed_frames}", flush=True)
                    print(f"[Worker] DEBUG Clip {clip_index}: next_idx in confirmed_frames = {next_idx in confirmed_frames}", flush=True)
                    
                    # Update start frame from previous odd clip's end
                    if prev_idx in confirmed_frames:
                        prev_end = confirmed_frames[prev_idx][1]
                        if prev_end:
                            old_start = clip_frames[clip_index]["start_frame"]
                            clip_frames[clip_index]["start_frame"] = prev_end
                            # Find index
                            for i, img in enumerate(images):
                                if img == prev_end or (hasattr(img, 'name') and hasattr(prev_end, 'name') and img.name == prev_end.name):
                                    clip_frames[clip_index]["start_index"] = i
                                    break
                            print(f"[Worker] Clip {clip_index}: Start frame updated {old_start.name if hasattr(old_start, 'name') else old_start} → {prev_end.name}", flush=True)
                    
                    # Update end frame from next odd clip's start
                    if next_idx in confirmed_frames:
                        next_start = confirmed_frames[next_idx][0]
                        if next_start:
                            old_end = clip_frames[clip_index].get("end_frame")
                            old_end_name = old_end.name if old_end and hasattr(old_end, 'name') else str(old_end)
                            clip_frames[clip_index]["end_frame"] = next_start
                            # Find index
                            for i, img in enumerate(images):
                                if img == next_start or (hasattr(img, 'name') and hasattr(next_start, 'name') and img.name == next_start.name):
                                    clip_frames[clip_index]["end_index"] = i
                                    break
                            print(f"[Worker] Clip {clip_index}: End frame updated {old_end_name} → {next_start.name}", flush=True)
                        else:
                            print(f"[Worker] DEBUG Clip {clip_index}: next_start is None/falsy!", flush=True)
                    elif clip_index == len(clip_frames) - 1:
                        # Last clip - no next odd clip, keep original end frame or use last frame
                        print(f"[Worker] Clip {clip_index}: Last clip, keeping assigned end frame", flush=True)
                    else:
                        print(f"[Worker] DEBUG Clip {clip_index}: next_idx {next_idx} NOT in confirmed_frames!", flush=True)
                    
                    # Debug: Print final frames for this clip
                    final_start = clip_frames[clip_index]["start_frame"]
                    final_end = clip_frames[clip_index].get("end_frame")
                    print(f"[Worker] DEBUG Clip {clip_index} FINAL: {final_start.name if hasattr(final_start, 'name') else final_start} → {final_end.name if final_end and hasattr(final_end, 'name') else final_end}", flush=True)
                    
                    # NOTE: Do NOT update clip.start_frame/end_frame here!
                    # They were set correctly at clip creation. The clip_frames values
                    # may be modified for CONTINUE mode chaining, but DB should preserve originals.
                
                # Log confirmed frames to job log for debugging
                with get_db() as db:
                    for idx, frames in confirmed_frames.items():
                        start_name = frames[0].name if frames[0] and hasattr(frames[0], 'name') else str(frames[0])
                        end_name = frames[1].name if frames[1] and hasattr(frames[1], 'name') else str(frames[1])
                        add_job_log(db, job_id, f"DEBUG: Clip {idx} confirmed frames: {start_name} → {end_name}", "DEBUG", "system")
                    
                    # Log the final frame assignments for even clips
                    for clip_index in even_indices:
                        final_start = clip_frames[clip_index]["start_frame"]
                        final_end = clip_frames[clip_index].get("end_frame")
                        start_name = final_start.name if hasattr(final_start, 'name') else str(final_start)
                        end_name = final_end.name if final_end and hasattr(final_end, 'name') else str(final_end)
                        add_job_log(db, job_id, f"DEBUG: Even clip {clip_index} will generate: {start_name} → {end_name}", "DEBUG", "system")
                
                with get_db() as db:
                    add_job_log(db, job_id, f"Phase 2: Processing even clips {[i+1 for i in even_indices]} with confirmed frames", "INFO", "system")
                
                # Process even clips in parallel
                with ThreadPoolExecutor(max_workers=min(parallel_clips, len(even_indices))) as executor:
                    # Phase 2: frames_locked=True - frames confirmed from Phase 1, cannot be swapped
                    futures = {executor.submit(process_clip_for_staggered, i, True): i for i in even_indices}
                    
                    for future in as_completed(futures):
                        clip_index = futures[future]
                        try:
                            result = future.result()
                            if result.get("success"):
                                completed += 1
                                if result.get("result", {}).get("output_path"):
                                    completed_clip_videos[clip_index] = str(result["result"]["output_path"])
                            elif result.get("skipped"):
                                skipped += 1
                                completed_clip_videos[clip_index] = None  # Mark as done for dependents
                                print(f"[Worker] Phase 2: Clip {clip_index} skipped, marking as done", flush=True)
                            else:
                                failed += 1
                            
                            # Update progress
                            with get_db() as db:
                                job = db.query(Job).filter(Job.id == job_id).first()
                                if job:
                                    job.completed_clips = completed
                                    job.failed_clips = failed
                                    job.skipped_clips = skipped
                                    processed = completed + failed + skipped
                                    job.progress_percent = (processed / total_clips) * 100 if total_clips > 0 else 0
                                    db.commit()
                        except Exception as e:
                            print(f"[Worker] Phase 2 future error for clip {clip_index}: {e}", flush=True)
                            failed += 1
            
            print(f"[Worker] === STAGGERED MODE COMPLETE ===", flush=True)
        
        # === PARALLEL MODE: Original batch processing ===
        else:
            print(f"[Worker] ⚡ PARALLEL MODE: Processing clips in parallel batches", flush=True)
            
            # Process clips with queue-based approach (ORIGINAL CODE)
            while (pending_clips or waiting_clips) and not generator.cancelled:
                # Check for redo clips and add them to pending
                redo_indices = check_redo_clips()
                if redo_indices:
                    for idx in redo_indices:
                        if idx not in pending_clips:
                            pending_clips.append(idx)
                    print(f"[Worker] Added {len(redo_indices)} redo clip(s) to pending queue", flush=True)
                
                # Check if keys are available before starting batch
                if not check_keys_available():
                    # Only log once per retry cycle
                    if no_keys_retries == 0:
                        print(f"[Worker] ⚠️ NO KEYS AVAILABLE - will pause job", flush=True)
                    no_keys_retries += 1
                    
                    if no_keys_retries > max_no_keys_retries:
                        # Max retries reached - PAUSE job instead of failing
                        with get_db() as db:
                            job = db.query(Job).filter(Job.id == job_id).first()
                            if job:
                                job.status = JobStatus.PAUSED.value
                                db.commit()
                            
                            add_job_log(
                                db, job_id,
                                f"⏸️ Job paused: API keys exhausted after {max_no_keys_retries} retries. Will auto-resume when keys available.",
                                "WARNING", "system"
                            )
                        
                        self._broadcast_event(job_id, {
                            "type": "job_paused_no_keys",
                            "message": "Job paused - waiting for API keys",
                        })
                        
                        generator.paused = True
                        raise JobPausedException("API keys exhausted - job paused")
                    
                    # Alert and wait (only once per retry)
                    send_no_keys_alert(job_id, no_keys_retries)
                    
                    # Wait with periodic checks (allow cancellation)
                    wait_end = time.time() + no_keys_wait_seconds
                    while time.time() < wait_end and not generator.cancelled:
                        if check_keys_available():
                            print(f"[Worker] ✅ Keys available again, resuming...", flush=True)
                            with get_db() as db:
                                add_job_log(db, job_id, "✅ API keys available, resuming generation", "INFO", "system")
                            break
                        time.sleep(10)  # Check every 10 seconds
                    
                    continue  # Re-check keys at top of loop
                
                # Reset retry counter when keys are available
                no_keys_retries = 0
                
                # Determine batch size based on available keys (using KeyPoolManager)
                from config import key_pool
                pool_status = key_pool.get_pool_status_summary(generator.api_keys)
                available_keys = pool_status["available"]
                total_keys = pool_status["total"]
                
                # Check for critical alerts only (no_keys) - low_keys is skipped
                if available_keys == 0:
                    send_key_alert_email("no_keys", available_keys, total_keys, job_id)
                elif available_keys > 15:
                    # Keys have recovered - reset alert flags
                    if _alerts_sent.get("no_keys"):
                        reset_key_alerts()
                
                # For continue mode clips in waiting_clips, check if previous clip is APPROVED
                # Also handle cases where previous clip was FAILED or SKIPPED
                newly_ready = []
                still_waiting = []
                clips_to_skip = []  # Clips whose predecessor was skipped/failed
                
                for clip_idx in waiting_clips:
                    prev_idx = clip_idx - 1
                    if prev_idx in approved_clip_videos:
                        newly_ready.append(clip_idx)
                        with get_db() as db:
                            clip = db.query(Clip).filter(
                                Clip.job_id == job_id,
                                Clip.clip_index == clip_idx
                            ).first()
                            if clip and clip.status == ClipStatus.WAITING_APPROVAL.value:
                                clip.status = ClipStatus.PENDING.value
                                db.commit()
                                print(f"[Worker] Clip {clip_idx}: Previous approved, moved to PENDING", flush=True)
                    elif prev_idx in completed_clip_videos and completed_clip_videos[prev_idx] is None:
                        # Previous clip completed but with no video - check actual DB status
                        # It might be in redo or still generating
                        with get_db() as db:
                            prev_clip = db.query(Clip).filter(
                                Clip.job_id == job_id,
                                Clip.clip_index == prev_idx
                            ).first()
                            if prev_clip:
                                # If previous clip is still being processed (redo, generating, pending), keep waiting
                                if prev_clip.status in [ClipStatus.GENERATING.value, ClipStatus.PENDING.value, 
                                                        ClipStatus.REDO_QUEUED.value, ClipStatus.COMPLETED.value]:
                                    still_waiting.append(clip_idx)
                                    print(f"[Worker] Clip {clip_idx}: Previous clip {prev_idx} status={prev_clip.status}, still waiting", flush=True)
                                else:
                                    # Truly failed or skipped
                                    clips_to_skip.append(clip_idx)
                            else:
                                clips_to_skip.append(clip_idx)
                    else:
                        still_waiting.append(clip_idx)
                
                # Handle clips whose predecessor was skipped/failed
                if clips_to_skip:
                    with get_db() as db:
                        for clip_idx in clips_to_skip:
                            prev_idx = clip_idx - 1
                            prev_clip = db.query(Clip).filter(
                                Clip.job_id == job_id,
                                Clip.clip_index == prev_idx
                            ).first()
                            
                            # Safety check: if prev_clip is now processing, skip this
                            if prev_clip and prev_clip.status in [
                                ClipStatus.GENERATING.value, ClipStatus.PENDING.value,
                                ClipStatus.REDO_QUEUED.value, ClipStatus.COMPLETED.value
                            ]:
                                print(f"[Worker] Clip {clip_idx}: Previous clip {prev_idx} now status={prev_clip.status}, skipping failure mark", flush=True)
                                continue
                            
                            clip = db.query(Clip).filter(
                                Clip.job_id == job_id,
                                Clip.clip_index == clip_idx
                            ).first()
                            
                            if clip and clip.status == ClipStatus.WAITING_APPROVAL.value:
                                if prev_clip and prev_clip.status == ClipStatus.SKIPPED.value:
                                    clip.status = ClipStatus.SKIPPED.value
                                    clip.error_code = "PREVIOUS_CLIP_SKIPPED"
                                    clip.error_message = f"Skipped: previous clip {prev_idx} was skipped"
                                    skipped += 1
                                    print(f"[Worker] Clip {clip_idx}: Previous clip {prev_idx} SKIPPED, marking as skipped", flush=True)
                                elif prev_clip and prev_clip.status == ClipStatus.FAILED.value:
                                    clip.status = ClipStatus.FAILED.value
                                    clip.error_code = "PREVIOUS_CLIP_FAILED"
                                    clip.error_message = f"Cannot process: previous clip {prev_idx} failed"
                                    failed += 1
                                    print(f"[Worker] Clip {clip_idx}: Previous clip {prev_idx} FAILED, marking as failed", flush=True)
                        db.commit()
                        
                        # Update job progress
                        job = db.query(Job).filter(Job.id == job_id).first()
                        if job:
                            job.completed_clips = completed
                            job.failed_clips = failed
                            job.skipped_clips = skipped
                            processed = completed + failed + skipped
                            job.progress_percent = (processed / total_clips) * 100 if total_clips > 0 else 0
                            db.commit()
                
                waiting_clips = still_waiting
                
                if newly_ready:
                    pending_clips.extend(newly_ready)
                    print(f"[Worker] {len(newly_ready)} clips now ready after approval", flush=True)
                
                ready_clips = pending_clips.copy()
                
                if not ready_clips:
                    # No clips ready - check if we're waiting for approvals
                    if waiting_clips:
                        # Still have clips waiting for approval - pause job processing
                        print(f"[Worker] {len(waiting_clips)} clips waiting for user approval", flush=True)
                        time.sleep(2)  # Check every 2 seconds for approvals
                        
                        # Check database for any approved OR FAILED clips
                        clips_to_remove = []
                        with get_db() as db:
                            for clip_idx in waiting_clips:
                                prev_idx = clip_idx - 1
                                prev_clip = db.query(Clip).filter(
                                    Clip.job_id == job_id,
                                    Clip.clip_index == prev_idx
                                ).first()
                                
                                if prev_clip:
                                    # Skip check if previous clip is still being processed
                                    if prev_clip.status in [ClipStatus.GENERATING.value, ClipStatus.PENDING.value,
                                                            ClipStatus.REDO_QUEUED.value]:
                                        # Previous clip still processing, keep waiting
                                        continue
                                    if prev_clip.approval_status == "approved":
                                        # Found an approval! Add to approved_clip_videos
                                        if prev_idx not in approved_clip_videos:
                                            video_path = None
                                            if prev_clip.output_filename:
                                                video_path = str(output_dir / prev_clip.output_filename)
                                            approved_clip_videos[prev_idx] = video_path
                                            print(f"[Worker] Detected approval for clip {prev_idx}, video_path={video_path}", flush=True)
                                    elif prev_clip.status == ClipStatus.FAILED.value:
                                        # Previous clip failed - this waiting clip should also fail or proceed without dependency
                                        print(f"[Worker] Clip {clip_idx}: Previous clip {prev_idx} FAILED, marking as failed", flush=True)
                                        clips_to_remove.append(clip_idx)
                                        
                                        # Mark this clip as failed too
                                        clip = db.query(Clip).filter(
                                            Clip.job_id == job_id,
                                            Clip.clip_index == clip_idx
                                        ).first()
                                        if clip and clip.status == ClipStatus.WAITING_APPROVAL.value:
                                            clip.status = ClipStatus.FAILED.value
                                            clip.error_code = "PREVIOUS_CLIP_FAILED"
                                            clip.error_message = f"Cannot process: previous clip {prev_idx} failed"
                                            db.commit()
                                            failed += 1
                                    elif prev_clip.status == ClipStatus.SKIPPED.value:
                                        # Previous clip was skipped (e.g., celebrity filter) - skip this one too
                                        print(f"[Worker] Clip {clip_idx}: Previous clip {prev_idx} SKIPPED, marking as skipped", flush=True)
                                        clips_to_remove.append(clip_idx)
                                        
                                        # Mark this clip as skipped too
                                        clip = db.query(Clip).filter(
                                            Clip.job_id == job_id,
                                            Clip.clip_index == clip_idx
                                        ).first()
                                        if clip and clip.status == ClipStatus.WAITING_APPROVAL.value:
                                            clip.status = ClipStatus.SKIPPED.value
                                            clip.error_code = "PREVIOUS_CLIP_SKIPPED"
                                            clip.error_message = f"Skipped: previous clip {prev_idx} was skipped"
                                            db.commit()
                                            skipped += 1
                        
                        # Remove clips whose dependency failed
                        for clip_idx in clips_to_remove:
                            if clip_idx in waiting_clips:
                                waiting_clips.remove(clip_idx)
                        
                        # Update job progress if any clips were failed
                        if clips_to_remove:
                            with get_db() as db:
                                job = db.query(Job).filter(Job.id == job_id).first()
                                if job:
                                    job.completed_clips = completed
                                    job.failed_clips = failed
                                    job.skipped_clips = skipped
                                    processed = completed + failed + skipped
                                    job.progress_percent = (processed / total_clips) * 100 if total_clips > 0 else 0
                                    db.commit()
                        
                        # Also check for redo clips during wait - process them immediately
                        redo_indices = check_redo_clips()
                        if redo_indices:
                            for idx in redo_indices:
                                if idx not in pending_clips:
                                    pending_clips.append(idx)
                            print(f"[Worker] Added {len(redo_indices)} redo clip(s) during approval wait", flush=True)
                        
                        continue
                    else:
                        # Nothing pending and nothing waiting - check one more time for redo clips
                        redo_indices = check_redo_clips()
                        if redo_indices:
                            for idx in redo_indices:
                                if idx not in pending_clips:
                                    pending_clips.append(idx)
                            print(f"[Worker] Added {len(redo_indices)} redo clip(s), continuing processing", flush=True)
                            continue
                        # Still nothing - we're done
                        break
                
                batch_size = min(parallel_clips, available_keys, len(ready_clips))
                
                if batch_size == 0:
                    continue
                
                batch = ready_clips[:batch_size]
                pending_clips = [c for c in pending_clips if c not in batch]
                
                print(f"[Worker] Processing batch of {batch_size} clips ({available_keys} keys available)", flush=True)
                print(f"[Worker] Batch clip indices: {batch}", flush=True)
                
                # Process batch in parallel
                with ThreadPoolExecutor(max_workers=parallel_clips) as clip_executor:
                    # Track active futures
                    futures = {}
                    for clip_idx in batch:
                        print(f"[Worker] Submitting clip {clip_idx} to executor...", flush=True)
                        future = clip_executor.submit(process_single_clip, clip_idx)
                        futures[future] = clip_idx
                        print(f"[Worker] Clip {clip_idx} submitted successfully", flush=True)
                    
                    active_count = len(futures)
                    
                    requeue_clips = []
                    
                    while futures and not generator.cancelled:
                        # Wait for at least one to complete (with timeout to check for new clips)
                        done_futures = set()
                        for future in list(futures.keys()):
                            if future.done():
                                done_futures.add(future)
                        
                        if not done_futures:
                            # No futures done yet, sleep briefly and check for new ready clips
                            time.sleep(0.5)
                            
                            # Check for redo clips while waiting
                            redo_indices = check_redo_clips()
                            if redo_indices:
                                for idx in redo_indices:
                                    if idx not in pending_clips and idx not in [futures[f] for f in futures]:
                                        pending_clips.append(idx)
                                if redo_indices:
                                    print(f"[Worker] Added {len(redo_indices)} redo clip(s) while processing batch", flush=True)
                            
                            # Check for newly approved clips
                            newly_ready_in_batch = []
                            still_waiting_in_batch = []
                            clips_to_skip_in_batch = []
                            for clip_idx in waiting_clips:
                                prev_idx = clip_idx - 1
                                if prev_idx in approved_clip_videos:
                                    newly_ready_in_batch.append(clip_idx)
                                    with get_db() as db:
                                        clip = db.query(Clip).filter(
                                            Clip.job_id == job_id,
                                            Clip.clip_index == clip_idx
                                        ).first()
                                        if clip and clip.status == ClipStatus.WAITING_APPROVAL.value:
                                            clip.status = ClipStatus.PENDING.value
                                            db.commit()
                                            print(f"[Worker] Clip {clip_idx}: Previous approved, moved to PENDING (during batch)", flush=True)
                                elif prev_idx in completed_clip_videos and completed_clip_videos[prev_idx] is None:
                                    # Previous clip completed but with no video - check actual DB status
                                    with get_db() as db:
                                        prev_clip_check = db.query(Clip).filter(
                                            Clip.job_id == job_id,
                                            Clip.clip_index == prev_idx
                                        ).first()
                                        if prev_clip_check and prev_clip_check.status in [
                                            ClipStatus.GENERATING.value, ClipStatus.PENDING.value,
                                            ClipStatus.REDO_QUEUED.value, ClipStatus.COMPLETED.value
                                        ]:
                                            still_waiting_in_batch.append(clip_idx)
                                            print(f"[Worker] Clip {clip_idx}: Previous clip {prev_idx} status={prev_clip_check.status}, still waiting", flush=True)
                                        else:
                                            clips_to_skip_in_batch.append(clip_idx)
                                else:
                                    # Also check database for approvals
                                    with get_db() as db:
                                        prev_clip = db.query(Clip).filter(
                                            Clip.job_id == job_id,
                                            Clip.clip_index == prev_idx
                                        ).first()
                                        if prev_clip and prev_clip.approval_status == "approved":
                                            if prev_idx not in approved_clip_videos:
                                                video_path = None
                                                if prev_clip.output_filename:
                                                    video_path = str(output_dir / prev_clip.output_filename)
                                                approved_clip_videos[prev_idx] = video_path
                                                newly_ready_in_batch.append(clip_idx)
                                                print(f"[Worker] Detected approval for clip {prev_idx} during batch, video_path={video_path}", flush=True)
                                        elif prev_clip and prev_clip.status in [ClipStatus.SKIPPED.value, ClipStatus.FAILED.value]:
                                            clips_to_skip_in_batch.append(clip_idx)
                                        else:
                                            still_waiting_in_batch.append(clip_idx)
                            
                            # Handle clips whose predecessor was skipped/failed during batch
                            if clips_to_skip_in_batch:
                                with get_db() as db:
                                    for clip_idx in clips_to_skip_in_batch:
                                        prev_idx = clip_idx - 1
                                        prev_clip = db.query(Clip).filter(
                                            Clip.job_id == job_id,
                                            Clip.clip_index == prev_idx
                                        ).first()
                                        
                                        # Safety check: if prev_clip is now processing, skip this
                                        if prev_clip and prev_clip.status in [
                                            ClipStatus.GENERATING.value, ClipStatus.PENDING.value,
                                            ClipStatus.REDO_QUEUED.value, ClipStatus.COMPLETED.value
                                        ]:
                                            print(f"[Worker] Clip {clip_idx}: Previous clip {prev_idx} now status={prev_clip.status}, skipping failure mark", flush=True)
                                            continue
                                        
                                        clip = db.query(Clip).filter(
                                            Clip.job_id == job_id,
                                            Clip.clip_index == clip_idx
                                        ).first()
                                        
                                        if clip and clip.status == ClipStatus.WAITING_APPROVAL.value:
                                            if prev_clip and prev_clip.status == ClipStatus.SKIPPED.value:
                                                clip.status = ClipStatus.SKIPPED.value
                                                clip.error_code = "PREVIOUS_CLIP_SKIPPED"
                                                clip.error_message = f"Skipped: previous clip {prev_idx} was skipped"
                                                skipped += 1
                                                print(f"[Worker] Clip {clip_idx}: Previous clip {prev_idx} SKIPPED (during batch)", flush=True)
                                            elif prev_clip and prev_clip.status == ClipStatus.FAILED.value:
                                                clip.status = ClipStatus.FAILED.value
                                                clip.error_code = "PREVIOUS_CLIP_FAILED"
                                                clip.error_message = f"Cannot process: previous clip {prev_idx} failed"
                                                failed += 1
                                                print(f"[Worker] Clip {clip_idx}: Previous clip {prev_idx} FAILED (during batch)", flush=True)
                                    db.commit()
                                    
                                    # Update job progress
                                    job = db.query(Job).filter(Job.id == job_id).first()
                                    if job:
                                        job.completed_clips = completed
                                        job.failed_clips = failed
                                        job.skipped_clips = skipped
                                        processed = completed + failed + skipped
                                        job.progress_percent = (processed / total_clips) * 100 if total_clips > 0 else 0
                                        db.commit()
                            
                            waiting_clips = still_waiting_in_batch
                            
                            # Add newly ready clips to pending
                            for idx in newly_ready_in_batch:
                                if idx not in pending_clips and idx not in [futures[f] for f in futures]:
                                    pending_clips.append(idx)
                            
                            continue  # Check again for new ready clips
                        
                        # Submit new clips if we have capacity
                        current_active = len([f for f in futures if not f.done()])
                        available_slots = parallel_clips - current_active
                        
                        if available_slots > 0 and pending_clips:
                            new_batch = pending_clips[:available_slots]
                            pending_clips = [c for c in pending_clips if c not in new_batch]
                            
                            for clip_idx in new_batch:
                                future = clip_executor.submit(process_single_clip, clip_idx)
                                futures[future] = clip_idx
                                print(f"[Worker] Submitted clip {clip_idx} to fill available slot", flush=True)
                        
                        # Process completed futures (removed errant 'continue' that made this unreachable)
                        for future in done_futures:
                            clip_index = futures.pop(future)
                            try:
                                result = future.result()
                                
                                if result.get("no_keys"):
                                    # Check if we should auto-pause the job
                                    if result.get("should_pause"):
                                        print(f"[Worker] Clip {clip_index} triggered auto-pause (keys exhausted after retries)", flush=True)
                                        # Set job to paused state and log it
                                        with get_db() as pause_db:
                                            pause_job = pause_db.query(Job).filter(Job.id == job_id).first()
                                            if pause_job:
                                                pause_job.status = JobStatus.PAUSED.value
                                                pause_db.commit()
                                            add_job_log(
                                                pause_db, job_id,
                                                f"⏸️ Job paused: API keys exhausted. Resume when quota resets (~2-3 min).",
                                                "WARNING", "system"
                                            )
                                        # Re-queue this clip and signal pause
                                        requeue_clips.append(clip_index)
                                        # Set generator pause flag
                                        generator.paused = True
                                    else:
                                        # Re-queue this clip for later
                                        requeue_clips.append(clip_index)
                                        print(f"[Worker] Clip {clip_index} failed due to no keys, re-queuing", flush=True)
                                elif result.get("success"):
                                    completed += 1
                                    # Track completed video for "continue" mode
                                    inner_result = result.get("result", {})
                                    if inner_result.get("output_path"):
                                        completed_clip_videos[clip_index] = str(inner_result["output_path"])
                                        print(f"[Worker] Tracked completed video for clip {clip_index}: {inner_result['output_path'].name}", flush=True)
                                elif result.get("skipped"):
                                    skipped += 1
                                    # For skipped clips, mark as "done" so dependent clips can fall back
                                    completed_clip_videos[clip_index] = None
                                    print(f"[Worker] Clip {clip_index} skipped, marking as done for dependents", flush=True)
                                else:
                                    failed += 1
                                    # For failed clips, still mark as "done" so dependent clips can fall back
                                    completed_clip_videos[clip_index] = None
                                
                                # Update job progress
                                with get_db() as db:
                                    job = db.query(Job).filter(Job.id == job_id).first()
                                    if job:
                                        job.completed_clips = completed
                                        job.failed_clips = failed
                                        job.skipped_clips = skipped
                                        processed = completed + failed + skipped
                                        job.progress_percent = (processed / total_clips) * 100 if total_clips > 0 else 0
                                        db.commit()
                                
                            except Exception as e:
                                print(f"[Worker] Future error for clip {clip_index}: {e}")
                                failed += 1
                                # Mark as done so dependents can proceed
                                completed_clip_videos[clip_index] = None
                                
                                # Update job progress after exception too
                                with get_db() as db:
                                    job = db.query(Job).filter(Job.id == job_id).first()
                                    if job:
                                        job.completed_clips = completed
                                        job.failed_clips = failed
                                        job.skipped_clips = skipped
                                        processed = completed + failed + skipped
                                        job.progress_percent = (processed / total_clips) * 100 if total_clips > 0 else 0
                                        db.commit()
                        
                        # === CHECK WAITING CLIPS AFTER PROCESSING FUTURES ===
                        # This is critical: when clips are skipped/failed, dependent clips need to be handled
                        if waiting_clips:
                            still_waiting_after = []
                            for clip_idx in waiting_clips:
                                prev_idx = clip_idx - 1
                                if prev_idx in completed_clip_videos and completed_clip_videos[prev_idx] is None:
                                    # Previous clip completed but with no video - check actual DB status
                                    with get_db() as db:
                                        prev_clip = db.query(Clip).filter(
                                            Clip.job_id == job_id,
                                            Clip.clip_index == prev_idx
                                        ).first()
                                        
                                        # If previous clip is still being processed, keep waiting
                                        if prev_clip and prev_clip.status in [
                                            ClipStatus.GENERATING.value, ClipStatus.PENDING.value,
                                            ClipStatus.REDO_QUEUED.value, ClipStatus.COMPLETED.value
                                        ]:
                                            still_waiting_after.append(clip_idx)
                                            print(f"[Worker] Clip {clip_idx}: Previous clip {prev_idx} status={prev_clip.status}, still waiting (after future)", flush=True)
                                            continue
                                        
                                        clip = db.query(Clip).filter(
                                            Clip.job_id == job_id,
                                            Clip.clip_index == clip_idx
                                        ).first()
                                        
                                        if clip and clip.status == ClipStatus.WAITING_APPROVAL.value:
                                            if prev_clip and prev_clip.status == ClipStatus.SKIPPED.value:
                                                clip.status = ClipStatus.SKIPPED.value
                                                clip.error_code = "PREVIOUS_CLIP_SKIPPED"
                                                clip.error_message = f"Skipped: previous clip {prev_idx} was skipped"
                                                skipped += 1
                                                print(f"[Worker] Clip {clip_idx}: Previous clip {prev_idx} SKIPPED (after future processing)", flush=True)
                                            elif prev_clip and prev_clip.status == ClipStatus.FAILED.value:
                                                clip.status = ClipStatus.FAILED.value
                                                clip.error_code = "PREVIOUS_CLIP_FAILED"
                                                clip.error_message = f"Cannot process: previous clip {prev_idx} failed"
                                                failed += 1
                                                print(f"[Worker] Clip {clip_idx}: Previous clip {prev_idx} FAILED (after future processing)", flush=True)
                                            db.commit()
                                            
                                            # Update job progress
                                            job = db.query(Job).filter(Job.id == job_id).first()
                                            if job:
                                                job.completed_clips = completed
                                                job.failed_clips = failed
                                                job.skipped_clips = skipped
                                                processed = completed + failed + skipped
                                                job.progress_percent = (processed / total_clips) * 100 if total_clips > 0 else 0
                                                db.commit()
                                else:
                                    still_waiting_after.append(clip_idx)
                            waiting_clips = still_waiting_after
                
                # Add re-queued clips back to pending
                if requeue_clips:
                    pending_clips = requeue_clips + pending_clips
                    print(f"[Worker] Re-queued {len(requeue_clips)} clips, {len(pending_clips)} pending", flush=True)
        
        # === APPROVAL WAIT LOOP ===
        # Job stays alive until ALL clips are approved (or job is cancelled)
        # This allows redos at any point before final approval
        print(f"[Worker] All clips processed. Waiting for approvals...", flush=True)
        
        with get_db() as db:
            add_job_log(db, job_id, "✅ All clips generated. Waiting for review & approval.", "INFO", "system")
        
        # Broadcast that we're now waiting for approvals
        self._broadcast_event(job_id, {
            "type": "awaiting_approval",
            "message": "All clips generated. Review and approve to complete job.",
        })
        
        last_status_log = time.time()
        
        # Track clips currently being processed in this loop
        processing_clips = set()
        processing_lock = threading.Lock()
        
        # Track which redo clips we've already logged about (to prevent spam)
        logged_redo_clips = set()
        
        def process_clip_async(clip_index: int, is_redo: bool = False):
            """Process a single clip (redo or newly-pending) asynchronously"""
            try:
                frames = clip_frames[clip_index]
                line_data = dialogue_data[clip_index]
                
                with get_db() as db:
                    clip = db.query(Clip).filter(
                        Clip.job_id == job_id,
                        Clip.clip_index == clip_index
                    ).first()
                    redo_feedback = clip.redo_feedback if clip and is_redo else None
                    
                    if clip:
                        clip.status = ClipStatus.GENERATING.value
                        clip.started_at = datetime.utcnow()
                        db.commit()
                        
                        # Get pool status for logging
                        from config import key_pool
                        pool_status = key_pool.get_pool_status_summary(generator.api_keys)
                        
                        if is_redo:
                            add_job_log(db, job_id, f"🔄 Processing redo for clip {clip_index + 1} (🔑 {pool_status['available']} keys working, {pool_status['rate_limited']} rate-limited)", "INFO", "redo")
                        else:
                            add_job_log(db, job_id, f"▶️ Processing clip {clip_index + 1} (predecessor approved) (🔑 {pool_status['available']} keys working, {pool_status['rate_limited']} rate-limited)", "INFO", "approval")
                
                self._broadcast_event(job_id, {
                    "type": "clip_started",
                    "clip_index": clip_index,
                    "start_frame": frames["start_frame"].name if hasattr(frames["start_frame"], 'name') else str(frames["start_frame"]),
                    "end_frame": frames["end_frame"].name if frames["end_frame"] and hasattr(frames["end_frame"], 'name') else None,
                })
                
                # CONTINUE mode: Extract frame from previous clip's video
                actual_start_frame = frames["start_frame"]
                original_scene_image = frames["start_frame"]  # Keep original for scene_image param
                clip_mode = frames.get("clip_mode", "blend")
                requires_previous = frames.get("requires_previous", False)
                
                if clip_mode == "continue" and requires_previous and clip_index > 0:
                    prev_idx = clip_index - 1
                    print(f"[Worker] process_clip_async: Clip {clip_index} is CONTINUE mode, checking for previous clip video", flush=True)
                    
                    # Look up previous clip's video from approved_clip_videos OR from database
                    prev_video = approved_clip_videos.get(prev_idx)
                    
                    if not prev_video:
                        # Try to get from database
                        with get_db() as db:
                            prev_clip = db.query(Clip).filter(
                                Clip.job_id == job_id,
                                Clip.clip_index == prev_idx
                            ).first()
                            if prev_clip and prev_clip.approval_status == "approved" and prev_clip.output_filename:
                                prev_video = str(output_dir / prev_clip.output_filename)
                                approved_clip_videos[prev_idx] = prev_video  # Cache it
                                print(f"[Worker] process_clip_async: Got previous video from DB: {prev_video}", flush=True)
                    
                    if prev_video and Path(prev_video).exists():
                        print(f"[Worker] process_clip_async: Extracting frame from {prev_video}", flush=True)
                        extracted = extract_frame_from_video(Path(prev_video), frame_offset=-8)
                        if extracted:
                            # Enhance with Nano Banana Pro
                            enhanced = enhance_frame_with_nano_banana(extracted, original_scene_image)
                            actual_start_frame = enhanced
                            print(f"[Worker] process_clip_async: Using {'enhanced' if enhanced != extracted else 'extracted'} frame", flush=True)
                        else:
                            print(f"[Worker] process_clip_async: Frame extraction failed, using original", flush=True)
                    else:
                        print(f"[Worker] process_clip_async: Previous video not available (prev_video={prev_video}), using original", flush=True)
                
                try:
                    result = generator.generate_single_clip(
                        start_frame=actual_start_frame,
                        end_frame=frames["end_frame"],
                        dialogue_line=line_data["text"],
                        dialogue_id=line_data["id"],
                        clip_index=clip_index,
                        output_dir=output_dir,
                        images_list=images,
                        current_end_index=frames["end_index"] if frames["end_index"] is not None else frames["start_index"],
                        scene_image=original_scene_image,  # Pass original for voice/analysis
                        redo_feedback=redo_feedback,
                        generation_mode=generation_mode,
                    )
                except Exception as e:
                    print(f"[Worker] Clip {clip_index} processing error: {e}", flush=True)
                    result = {"success": False, "error": str(e)}
                
                # Update clip record
                nonlocal completed, failed
                with get_db() as db:
                    clip = db.query(Clip).filter(
                        Clip.job_id == job_id,
                        Clip.clip_index == clip_index
                    ).first()
                    if clip:
                        clip.completed_at = datetime.utcnow()
                        if is_redo:
                            clip.redo_feedback = None
                        if result.get("success"):
                            new_filename = result["output_path"].name if result.get("output_path") else None
                            
                            # Update versions_json properly
                            versions = json.loads(clip.versions_json) if clip.versions_json else []
                            existing_attempts = [v.get('attempt') for v in versions]
                            current_attempt = clip.generation_attempt or 1
                            
                            if current_attempt not in existing_attempts and new_filename:
                                versions.append({
                                    "attempt": current_attempt,
                                    "filename": new_filename,
                                    "generated_at": datetime.utcnow().isoformat(),
                                })
                                clip.versions_json = json.dumps(versions)
                            
                            clip.status = ClipStatus.COMPLETED.value
                            clip.output_filename = new_filename
                            clip.selected_variant = len(versions)
                            clip.approval_status = "pending_review"
                            completed += 1
                            if result.get("output_path"):
                                video_path = str(result["output_path"])
                                completed_clip_videos[clip_index] = video_path
                                # NOTE: Don't add to approved_clip_videos here!
                                # CONTINUE mode clips must wait for user approval first.
                                # approved_clip_videos is populated when approval is detected in waiting_clips check.
                                
                                # Upload to R2 for persistence (API jobs)
                                try:
                                    from backends.storage import is_storage_configured, get_storage
                                    if is_storage_configured():
                                        storage = get_storage()
                                        r2_key = f"jobs/{job_id}/outputs/{new_filename}"
                                        storage.upload_file(video_path, r2_key, content_type='video/mp4')
                                        output_url = storage.get_presigned_url(r2_key, expires_in=86400 * 7)
                                        clip.output_url = output_url
                                        # Update version entry with URL
                                        if versions:
                                            versions[-1]["url"] = output_url
                                            clip.versions_json = json.dumps(versions)
                                        print(f"[Worker] Uploaded clip {clip_index} to R2: {r2_key}", flush=True)
                                except Exception as r2_err:
                                    print(f"[Worker] R2 upload failed for clip {clip_index} (non-fatal): {r2_err}", flush=True)
                        else:
                            # Check if this is a "no keys" situation - re-queue as redo
                            if result.get("no_keys") or result.get("should_pause"):
                                # Re-queue as redo to be picked up when keys are available
                                clip.status = ClipStatus.REDO_QUEUED.value
                                add_job_log(
                                    db, job_id,
                                    f"Clip {clip_index + 1} re-queued: API keys temporarily unavailable",
                                    "WARNING", "system"
                                )
                            else:
                                clip.status = ClipStatus.FAILED.value
                                error_obj = result.get("error")
                                if error_obj:
                                    clip.error_message = str(error_obj)[:500]
                                failed += 1
                        db.commit()
                
                self._broadcast_event(job_id, {
                    "type": "clip_completed",
                    "clip_index": clip_index,
                    "success": result.get("success", False),
                    "output": result["output_path"].name if result.get("output_path") else None,
                })
                
            finally:
                with processing_lock:
                    processing_clips.discard(clip_index)
        
        # Create executor for parallel clip processing in approval loop
        from concurrent.futures import ThreadPoolExecutor
        approval_executor = ThreadPoolExecutor(max_workers=6)
        
        try:
            while not generator.cancelled:
                with get_db() as db:
                    # Count clips by status
                    clips = db.query(Clip).filter(Clip.job_id == job_id).all()
                    
                    total = len(clips)
                    approved_count = sum(1 for c in clips if c.approval_status == "approved")
                    pending_review_count = sum(1 for c in clips if c.approval_status == "pending_review")
                    redo_queued_count = sum(1 for c in clips if c.status == ClipStatus.REDO_QUEUED.value)
                    generating_count = sum(1 for c in clips if c.status == ClipStatus.GENERATING.value)
                    failed_count = sum(1 for c in clips if c.status == ClipStatus.FAILED.value)
                    skipped_count = sum(1 for c in clips if c.status == ClipStatus.SKIPPED.value)
                    waiting_approval_count = sum(1 for c in clips if c.status == ClipStatus.WAITING_APPROVAL.value)
                    pending_count = sum(1 for c in clips if c.status == ClipStatus.PENDING.value)
                    
                    # Handle stuck WAITING_APPROVAL clips whose predecessors are skipped/failed
                    if waiting_approval_count > 0:
                        for clip in clips:
                            if clip.status == ClipStatus.WAITING_APPROVAL.value:
                                prev_idx = clip.clip_index - 1
                                prev_clip = next((c for c in clips if c.clip_index == prev_idx), None)
                                if prev_clip:
                                    # Don't mark as failed if previous clip is still being processed
                                    if prev_clip.status in [ClipStatus.GENERATING.value, ClipStatus.PENDING.value,
                                                            ClipStatus.REDO_QUEUED.value, ClipStatus.COMPLETED.value]:
                                        # Still waiting for previous clip
                                        continue
                                    if prev_clip.status == ClipStatus.SKIPPED.value:
                                        clip.status = ClipStatus.SKIPPED.value
                                        clip.error_code = "PREVIOUS_CLIP_SKIPPED"
                                        clip.error_message = f"Skipped: previous clip {prev_idx} was skipped"
                                        skipped_count += 1
                                        print(f"[Worker] Clip {clip.clip_index}: Previous clip {prev_idx} SKIPPED (approval loop cleanup)", flush=True)
                                    elif prev_clip.status == ClipStatus.FAILED.value:
                                        clip.status = ClipStatus.FAILED.value
                                        clip.error_code = "PREVIOUS_CLIP_FAILED"
                                        clip.error_message = f"Cannot process: previous clip {prev_idx} failed"
                                        failed_count += 1
                                        print(f"[Worker] Clip {clip.clip_index}: Previous clip {prev_idx} FAILED (approval loop cleanup)", flush=True)
                        db.commit()
                    
                    # Check if all clips are approved (excluding failed and skipped ones - they can't be approved)
                    approvable_clips = total - failed_count - skipped_count
                    all_approved = (approved_count >= approvable_clips) and approvable_clips > 0
                    
                    # Also complete job if ALL clips are skipped/failed (nothing left to approve)
                    all_terminal = (failed_count + skipped_count >= total)
                    
                    if all_approved:
                        print(f"[Worker] All {approved_count} clips approved! Job complete.", flush=True)
                        add_job_log(db, job_id, f"🎉 All {approved_count} clips approved!", "INFO", "system")
                        break
                    
                    if all_terminal:
                        print(f"[Worker] All clips are terminal (skipped/failed). Job complete.", flush=True)
                        add_job_log(db, job_id, f"⚠️ Job complete: {skipped_count} skipped, {failed_count} failed. No clips to approve.", "WARNING", "system")
                        break
                    
                    # Find PENDING clips that need processing (from WAITING_APPROVAL transitions)
                    pending_indices = [c.clip_index for c in clips if c.status == ClipStatus.PENDING.value]
                
                # Check for redo requests (just get indices, don't change status)
                redo_indices = check_redo_clips()
                
                # Submit pending clips for processing (these waited for predecessor approval)
                with processing_lock:
                    for clip_index in pending_indices:
                        if clip_index not in processing_clips:
                            processing_clips.add(clip_index)
                            print(f"[Worker] Submitting pending clip {clip_index + 1} for processing (predecessor approved)", flush=True)
                            approval_executor.submit(process_clip_async, clip_index, False)
                
                # Note: Redos are handled by the independent _check_redo_queue() processor
                # We just log NEW redo requests here (avoid spam)
                if redo_indices:
                    new_redos = [i for i in redo_indices if i not in logged_redo_clips]
                    if new_redos:
                        print(f"[Worker] Redo requests detected: clips {[i+1 for i in new_redos]} (handled by independent processor)", flush=True)
                        logged_redo_clips.update(new_redos)
                    
                    # Clear logged clips that are no longer in redo queue (redo completed or cancelled)
                    logged_redo_clips.intersection_update(redo_indices)
            
            # Sleep before next check (1 second polling)
            time.sleep(1)
            
            # Log status every 30 seconds so user knows job is still active
            if time.time() - last_status_log > 30:
                last_status_log = time.time()
                with get_db() as db:
                    clips = db.query(Clip).filter(Clip.job_id == job_id).all()
                    approved = sum(1 for c in clips if c.approval_status == "approved")
                    pending = sum(1 for c in clips if c.approval_status == "pending_review")
                    redo_queued = [c for c in clips if c.status == ClipStatus.REDO_QUEUED.value]
                    total = len(clips)
                    print(f"[Worker] Approval status: {approved}/{total} approved, {pending} pending review", flush=True)
                    
                    # Warn about stuck redos
                    if redo_queued:
                        print(f"[Worker] ⚠️ {len(redo_queued)} clips stuck in REDO_QUEUED: {[c.clip_index + 1 for c in redo_queued]}", flush=True)
                        print(f"[Worker] _processing_redo_clips has {len(self._processing_redo_clips)} items", flush=True)
        
        finally:
            # Cleanup executor
            approval_executor.shutdown(wait=False)
        
        # Job completed - calculate status from actual clip data
        actual_completed = 0
        actual_failed = 0
        actual_skipped = 0
        final_status = "unknown"
        
        with get_db() as db:
            job = db.query(Job).filter(Job.id == job_id).first()
            if job:
                # Check if job was already cancelled by user - don't overwrite
                if job.status == JobStatus.CANCELLED.value:
                    print(f"[Worker] Job already cancelled by user, skipping status update", flush=True)
                    # Still update stats
                    clips = db.query(Clip).filter(Clip.job_id == job_id).all()
                    actual_completed = sum(1 for c in clips if c.status == ClipStatus.COMPLETED.value)
                    actual_failed = sum(1 for c in clips if c.status == ClipStatus.FAILED.value)
                    actual_skipped = sum(1 for c in clips if c.status == ClipStatus.SKIPPED.value)
                    job.completed_clips = actual_completed
                    job.failed_clips = actual_failed
                    job.skipped_clips = actual_skipped
                    db.commit()
                    final_status = job.status
                else:
                    # Recalculate stats from actual clips in database
                    clips = db.query(Clip).filter(Clip.job_id == job_id).all()
                    actual_completed = sum(1 for c in clips if c.status == ClipStatus.COMPLETED.value)
                    actual_failed = sum(1 for c in clips if c.status == ClipStatus.FAILED.value)
                    actual_skipped = sum(1 for c in clips if c.status == ClipStatus.SKIPPED.value)
                    
                    job.completed_clips = actual_completed
                    job.failed_clips = actual_failed
                    job.skipped_clips = actual_skipped
                    job.progress_percent = 100.0
                    
                    if generator.cancelled:
                        job.status = JobStatus.CANCELLED.value
                    elif actual_completed == 0 and actual_failed > 0:
                        job.status = JobStatus.FAILED.value
                    else:
                        job.status = JobStatus.COMPLETED.value
                    
                    job.completed_at = datetime.utcnow()
                    final_status = job.status
                    db.commit()
                    
                    add_job_log(
                        db, job_id,
                        f"Job completed: {actual_completed} success, {actual_failed} failed, {actual_skipped} skipped",
                        "INFO", "system"
                    )
                    
                    # Generate missing clips file for celebrity-filtered clips
                    celebrity_skipped = [c for c in clips if c.status == ClipStatus.SKIPPED.value and c.error_code == "CELEBRITY_FILTER"]
                    if celebrity_skipped:
                        # Generate Excel file with missing clips info
                        missing_clips_path = Path(job.output_dir) / "missing_clips.xlsx"
                        try:
                            # Try to use openpyxl for Excel
                            try:
                                from openpyxl import Workbook
                                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                                
                                wb = Workbook()
                                ws = wb.active
                                ws.title = "Missing Clips"
                                
                                # Header style
                                header_font = Font(bold=True, color="FFFFFF")
                                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                                thin_border = Border(
                                    left=Side(style='thin'),
                                    right=Side(style='thin'),
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin')
                                )
                                
                                # Headers
                                headers = ["Clip #", "Start Image", "End Image", "Dialogue", "Prompt"]
                                for col, header in enumerate(headers, 1):
                                    cell = ws.cell(row=1, column=col, value=header)
                                    cell.font = header_font
                                    cell.fill = header_fill
                                    cell.alignment = Alignment(horizontal='center', vertical='center')
                                    cell.border = thin_border
                                
                                # Data rows
                                for row, clip in enumerate(celebrity_skipped, 2):
                                    ws.cell(row=row, column=1, value=clip.clip_index + 1).border = thin_border
                                    ws.cell(row=row, column=2, value=clip.start_frame or "").border = thin_border
                                    ws.cell(row=row, column=3, value=clip.end_frame or "").border = thin_border
                                    ws.cell(row=row, column=4, value=clip.dialogue_text or "").border = thin_border
                                    ws.cell(row=row, column=5, value=clip.prompt_text or "").border = thin_border
                                    
                                    # Wrap text for dialogue and prompt columns
                                    ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical='top')
                                    ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical='top')
                                
                                # Set column widths
                                ws.column_dimensions['A'].width = 8   # Clip #
                                ws.column_dimensions['B'].width = 20  # Start Image
                                ws.column_dimensions['C'].width = 20  # End Image
                                ws.column_dimensions['D'].width = 50  # Dialogue
                                ws.column_dimensions['E'].width = 80  # Prompt
                                
                                # Add note at the top
                                ws.insert_rows(1)
                                ws.merge_cells('A1:E1')
                                note_cell = ws.cell(row=1, column=1, value="⚠️ These clips were skipped due to celebrity filter. You can try generating them manually in Google AI Studio. Eligible for reimbursement.")
                                note_cell.font = Font(bold=True, color="FF6600")
                                note_cell.alignment = Alignment(horizontal='left', vertical='center')
                                ws.row_dimensions[1].height = 25
                                
                                wb.save(missing_clips_path)
                                
                            except ImportError:
                                # Fallback to CSV if openpyxl not available
                                import csv
                                missing_clips_path = Path(job.output_dir) / "missing_clips.csv"
                                with open(missing_clips_path, 'w', newline='', encoding='utf-8') as f:
                                    writer = csv.writer(f)
                                    writer.writerow(["Clip #", "Start Image", "End Image", "Dialogue", "Prompt"])
                                    for clip in celebrity_skipped:
                                        writer.writerow([
                                            clip.clip_index + 1,
                                            clip.start_frame or "",
                                            clip.end_frame or "",
                                            clip.dialogue_text or "",
                                            clip.prompt_text or ""
                                        ])
                            
                            add_job_log(
                                db, job_id,
                                f"📋 {len(celebrity_skipped)} clip(s) skipped due to celebrity filter. Details saved to {missing_clips_path.name}",
                                "WARNING", "celebrity_filter"
                            )
                        except Exception as e:
                            print(f"[Worker] Failed to save missing_clips file: {e}", flush=True)
            
            # Save blacklist
            for img_path in generator.blacklist:
                entry = BlacklistEntry(
                    job_id=job_id,
                    image_filename=img_path.name,
                    reason="generation_failed",
                )
                db.add(entry)
            db.commit()
        
        self._broadcast_event(job_id, {
            "type": "job_completed",
            "status": final_status,
            "completed": actual_completed,
            "failed": actual_failed,
            "skipped": actual_skipped,
        })
    
    def _get_next_clean_image(
        self,
        generator: VeoGenerator,
        images: List[Path],
        current_index: int,
    ) -> Optional[tuple]:
        """Get next non-blacklisted image"""
        total = len(images)
        
        for offset in range(1, min(generator.config.max_image_attempts + 1, total + 1)):
            new_index = (current_index + offset) % total
            candidate = images[new_index]
            
            if candidate not in generator.blacklist:
                return (new_index, candidate)
        
        return None
    
    def _get_next_clean_start(
        self,
        generator: VeoGenerator,
        images: List[Path],
        current_index: int,
    ) -> Optional[tuple]:
        """Get next non-blacklisted start frame"""
        return self._get_next_clean_image(generator, images, current_index)
    
    def _handle_progress(
        self,
        job_id: str,
        clip_index: int,
        status: str,
        message: str,
        details: Optional[Dict],
    ):
        """Handle progress update from generator"""
        with get_db() as db:
            add_job_log(
                db, job_id, message,
                level="INFO" if status != "error" else "ERROR",
                category="clip",
                clip_index=clip_index,
                details=details
            )
        
        self._broadcast_event(job_id, {
            "type": "progress",
            "clip_index": clip_index,
            "status": status,
            "message": message,
            "details": details,
        })
    
    def _handle_error(self, job_id: str, error: VeoError):
        """Handle error from generator"""
        with get_db() as db:
            add_job_log(
                db, job_id,
                error.message,
                level="ERROR",
                category="error",
                details=error.to_dict()
            )
        
        self._broadcast_event(job_id, {
            "type": "error",
            "error": error.to_dict(),
        })
    
    # ============ SSE Subscription Management ============
    
    def subscribe(self, job_id: str) -> Queue:
        """Subscribe to job events"""
        event_queue = Queue()
        
        with self.subscribers_lock:
            if job_id not in self.subscribers:
                self.subscribers[job_id] = []
            self.subscribers[job_id].append(event_queue)
            print(f"[Worker] Subscribed to job {job_id[:8]}, total subscribers: {len(self.subscribers[job_id])}", flush=True)
        
        return event_queue
    
    def unsubscribe(self, job_id: str, event_queue: Queue):
        """Unsubscribe from job events"""
        with self.subscribers_lock:
            if job_id in self.subscribers:
                if event_queue in self.subscribers[job_id]:
                    self.subscribers[job_id].remove(event_queue)
                    print(f"[Worker] Unsubscribed from job {job_id[:8]}, remaining: {len(self.subscribers[job_id])}", flush=True)
                if not self.subscribers[job_id]:
                    del self.subscribers[job_id]
    
    def _broadcast_event(self, job_id: str, event: Dict):
        """Broadcast event to all subscribers"""
        print(f"[Worker] Broadcasting event: {event.get('type')} for job {job_id[:8]}", flush=True)
        with self.subscribers_lock:
            if job_id in self.subscribers:
                subscriber_count = len(self.subscribers[job_id])
                print(f"[Worker] Broadcasting to {subscriber_count} subscribers", flush=True)
                for queue in self.subscribers[job_id]:
                    try:
                        queue.put_nowait(event)
                    except Exception as e:
                        print(f"[Worker] Failed to broadcast: {e}", flush=True)
            else:
                print(f"[Worker] No subscribers for job {job_id[:8]}", flush=True)
    
    # ============ Job Control ============
    
    def cancel_job(self, job_id: str) -> bool:
        """Cancel a running job"""
        cancelled = False
        completed_count = 0
        failed_count = 0
        skipped_count = 0
        
        # Cancel the generator if running
        if job_id in self.running_jobs:
            generator = self.running_jobs[job_id]
            if generator is not None:  # Could be None placeholder
                generator.cancel()
                cancelled = True
        
        # Always update job status in database
        with get_db() as db:
            job = db.query(Job).filter(Job.id == job_id).first()
            if job:
                # Calculate final stats from clips
                clips = db.query(Clip).filter(Clip.job_id == job_id).all()
                completed_count = sum(1 for c in clips if c.status == ClipStatus.COMPLETED.value)
                failed_count = sum(1 for c in clips if c.status == ClipStatus.FAILED.value)
                skipped_count = sum(1 for c in clips if c.status == ClipStatus.SKIPPED.value)
                
                # Only update if still running (not already completed/failed)
                if job.status == JobStatus.RUNNING.value:
                    job.status = JobStatus.CANCELLED.value
                    job.completed_clips = completed_count
                    job.failed_clips = failed_count
                    job.skipped_clips = skipped_count
                    job.completed_at = datetime.utcnow()
                    db.commit()
                    
                    add_job_log(db, job_id, f"Job cancelled by user: {completed_count} completed, {failed_count} failed", "INFO", "system")
                    cancelled = True
                else:
                    # Job already completed/failed, just log
                    print(f"[Worker] cancel_job called but job status is {job.status}, not updating", flush=True)
        
        # Broadcast cancellation event to UI (always if we tried to cancel)
        self._broadcast_event(job_id, {
            "type": "job_completed",
            "status": "cancelled",
            "completed": completed_count,
            "failed": failed_count,
            "skipped": skipped_count,
        })
        
        # Clean up running_jobs
        if job_id in self.running_jobs:
            del self.running_jobs[job_id]
        
        return cancelled
    
    def pause_job(self, job_id: str) -> bool:
        """Pause a running job"""
        if job_id in self.running_jobs:
            self.running_jobs[job_id].pause()
            
            with get_db() as db:
                job = db.query(Job).filter(Job.id == job_id).first()
                if job:
                    job.status = JobStatus.PAUSED.value
                    db.commit()
            
            return True
        return False
    
    def resume_job(self, job_id: str) -> bool:
        """Resume a paused job"""
        # Case 1: Job has active generator - just unpause it
        if job_id in self.running_jobs:
            self.running_jobs[job_id].resume()
            
            with get_db() as db:
                job = db.query(Job).filter(Job.id == job_id).first()
                if job:
                    job.status = JobStatus.RUNNING.value
                    db.commit()
            
            return True
        
        # Case 2: Job was paused before generator started - re-queue it
        with get_db() as db:
            job = db.query(Job).filter(Job.id == job_id).first()
            if job and job.status == JobStatus.PAUSED.value:
                # Set to pending so worker picks it up
                job.status = JobStatus.PENDING.value
                db.commit()
                
                # Add to queue
                self.job_queue.put(job_id)
                print(f"[Worker] Re-queued paused job {job_id[:8]} for processing", flush=True)
                return True
        
        return False
    
    def get_job_status(self, job_id: str) -> Optional[Dict]:
        """Get current job status"""
        with get_db() as db:
            job = db.query(Job).filter(Job.id == job_id).first()
            if job:
                return job.to_dict()
        return None


# Singleton worker instance
worker = JobWorker(max_workers=app_config.max_workers)
print("=" * 60, flush=True)
print(f"WORKER VERSION: {WORKER_VERSION}", flush=True)
print("=" * 60, flush=True)